"""Loader for the cleaned, atomic gold sheet (one row = one case).

Detects the atomic schema produced for manual review and builds GoldCorpusCase
objects directly, honoring the manual "Оставить (defect/opinion)" column. Works
for .xlsx and .csv. Falls back (returns None) when the sheet is the legacy
Проект/Проблема/Детали format, so callers keep using the old splitting loader.
"""

from __future__ import annotations

import csv as _csv
from pathlib import Path

from content_audit.corpus_evaluation import GoldCorpusCase, GoldCorpusItem, _match_project

CRIT_CODE = "критерий_код"
TEXT_COLS = ("Текст кейса", "Текст ошибки", "Текст")
KEEP_COLS = ("Оставить (defect/opinion)", "Оставить", "Тип (авто)", "Тип")
PROJECT_COL = "Проект"
LINE_COL = "Строка"


def _read_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig") as fh:
            return [list(r) for r in _csv.reader(fh)]
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]


def _first(header: dict, names) -> int | None:
    for n in names:
        if n in header:
            return header[n]
    return None


def _parse_line(value: str):
    v = (value or "").strip()
    if not v:
        return None, None
    import re

    m = re.match(r"^\s*(\d+)\s*[-–—]\s*(\d+)\s*$", v)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return min(a, b), max(a, b)
    m = re.match(r"^\s*(\d+)\s*$", v)
    if m:
        return int(m.group(1)), int(m.group(1))
    return None, None


def is_atomic(path: Path) -> bool:
    try:
        rows = _read_rows(path)
    except Exception:
        return False
    if not rows:
        return False
    header = {str(h).strip(): i for i, h in enumerate(rows[0]) if str(h).strip()}
    return CRIT_CODE in header and _first(header, TEXT_COLS) is not None


def load_atomic(path: Path, unit_candidates):
    """Returns (gold_items, gold_cases, opinion_case_ids, notes)."""

    rows = _read_rows(path)
    header = {str(h).strip(): i for i, h in enumerate(rows[0]) if str(h).strip()}
    ci_crit = header[CRIT_CODE]
    ci_text = _first(header, TEXT_COLS)
    ci_keep = _first(header, KEEP_COLS)
    ci_proj = header.get(PROJECT_COL)
    ci_line = header.get(LINE_COL)
    ci_file = header.get("Файл (якорь)") if "Файл (якорь)" in header else header.get("Файл")

    cases: list[GoldCorpusCase] = []
    items: list[GoldCorpusItem] = []
    opinion_ids: set[str] = set()
    notes: list[str] = []

    def cell(row, idx):
        return (row[idx].strip() if idx is not None and idx < len(row) and row[idx] is not None else "")

    for rownum, row in enumerate(rows[1:], start=2):
        crit = cell(row, ci_crit)
        text = cell(row, ci_text)
        if not crit or not text:
            continue
        raw_project = cell(row, ci_proj) if ci_proj is not None else ""
        candidate, score = _match_project(raw_project, unit_candidates)
        if score < 0.55:
            notes.append(
                "Строка %d: слабое сопоставление проекта %r с папкой %r, score=%.2f."
                % (rownum, raw_project, candidate.raw_name, score)
            )
        line_start, line_end = _parse_line(cell(row, ci_line)) if ci_line is not None else (None, None)
        file_hint = (cell(row, ci_file).split(";")[0].strip() or None) if ci_file is not None else None
        case_id = "gold_%d_%s" % (rownum, crit)
        case = GoldCorpusCase(
            case_id=case_id,
            row_number=rownum,
            raw_project=raw_project,
            matched_project=candidate.raw_name,
            project_id=candidate.project_id,
            criterion=crit,
            line_start=line_start,
            line_end=line_end,
            gold_text=text,
            file_hint=file_hint,
        )
        cases.append(case)
        items.append(
            GoldCorpusItem(
                row_number=rownum,
                raw_project=raw_project,
                matched_project=candidate.raw_name,
                project_id=candidate.project_id,
                raw_problem=crit,
                details=text,
                criteria=[crit],
            )
        )
        keep = cell(row, ci_keep).lower() if ci_keep is not None else ""
        if keep.startswith("opinion") or keep in ("мнение", "opinion", "wish", "пожелание"):
            opinion_ids.add(case_id)
    return items, cases, opinion_ids, notes
