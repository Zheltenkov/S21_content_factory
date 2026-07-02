from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl
from methodologist_review_resolution import apply_methodologist_review_decisions
from review_queue_messages import humanize_review_details


SKIP_SHEETS = {
    "пример",
    "структура",
    "доп",
    "дополнительно",
    "учесть",
    "лист10",
}

EXCLUDED_WORKBOOK_PATTERNS = (
    "шаблон компетентностного профиля",
)

LEVEL_HINTS = (
    "junior",
    "middle",
    "senior",
    "стажер",
    "начальн",
    "базов",
    "продвинут",
    "мастер",
    "уверенн",
)

DIMENSION_MAP = {
    "знает": ("knowledge", "Знает"),
    "умеет": ("ability", "Умеет"),
    "владеет": ("proficiency", "Владеет"),
    "понимает": ("understanding", "Понимает"),
}


@dataclass
class BlockHeader:
    header_row_idx: int
    level_row_idx: int
    title: str | None
    description: str | None
    prerequisites: str | None
    skill_col: int
    dimension_col: int
    number_col: int
    level_headers: list[tuple[int, str]]
    meta_headers: list[tuple[int, str]]


def normalize_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def normalize_key(value: object | None) -> str:
    text = normalize_text(value)
    return text.lower() if text else ""


def slugify(text: str) -> str:
    text = normalize_key(text)
    text = re.sub(r"[^\w]+", "-", text, flags=re.UNICODE)
    return text.strip("-") or "unnamed"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def detect_source_kind(file_name: str) -> str:
    lower = normalize_key(file_name)
    if lower.startswith("шаблон"):
        return "template"
    if lower.startswith("сырой"):
        return "draft"
    return "role_profile"


def clean_profile_name(file_name: str) -> str:
    name = Path(file_name).stem
    name = re.sub(r"^(NEW|new|v\d+|V\d+|сырой)\s+", "", name).strip()
    name = re.sub(r"^Компетентностный профиль\s*", "", name).strip()
    return name.strip(" -")


def workbook_paths(root: Path) -> list[Path]:
    files = []
    for path in root.rglob("*.xlsx"):
        if ".__" in path.name or path.name.startswith("._"):
            continue
        if "__MACOSX" in path.parts:
            continue
        normalized_name = normalize_key(path.name).replace("_", " ").replace("/", " ")
        if any(pattern in normalized_name for pattern in EXCLUDED_WORKBOOK_PATTERNS):
            continue
        if "ux" in normalized_name and "ui" in normalized_name and "дизайнер" in normalized_name:
            continue
        files.append(path)
    return sorted(files)


def is_skip_sheet(sheet_name: str) -> bool:
    return normalize_key(sheet_name) in SKIP_SHEETS


def is_header_row(row: list[object | None]) -> bool:
    keys = [normalize_key(cell) for cell in row]
    return bool(keys) and keys[0] in {"№", "n"} and "skills" in keys and "х" in keys


def row_has_any_content(row: Iterable[object | None]) -> bool:
    return any(normalize_text(cell) is not None for cell in row)


def is_level_header(text: str) -> bool:
    lowered = normalize_key(text)
    return any(token in lowered for token in LEVEL_HINTS)


def canonical_band(text: str) -> str | None:
    lowered = normalize_key(text)
    if "стаж" in lowered:
        return "trainee"
    if "junior-" in lowered:
        return "junior_minus"
    if "junior+" in lowered:
        return "junior_plus"
    if "junior" in lowered or "начальн" in lowered:
        return "junior"
    if "middle" in lowered or "уверенн" in lowered:
        return "middle"
    if "senior" in lowered:
        return "senior"
    if "мастер" in lowered:
        return "master"
    if lowered == "базовый":
        return "basic"
    if lowered == "начальный":
        return "beginner"
    return None


def row_signature(row: list[object | None]) -> list[str]:
    return [normalize_text(cell) for cell in row if normalize_text(cell)]


def find_block_starts(rows: list[list[object | None]]) -> list[int]:
    return [index for index, row in enumerate(rows) if is_header_row(row)]


def build_block_header(
    rows: list[list[object | None]],
    start_idx: int,
    previous_level_headers: list[tuple[int, str]] | None,
) -> tuple[BlockHeader, list[str]]:
    issues: list[str] = []
    header_row = rows[start_idx]
    level_row = rows[start_idx + 1] if start_idx + 1 < len(rows) else []

    skill_col = next(i for i, cell in enumerate(header_row) if normalize_key(cell) == "skills")
    dimension_col = next(i for i, cell in enumerate(header_row) if normalize_key(cell) == "х")
    number_col = 0

    title = description = prerequisites = None
    meta_candidates = rows[max(0, start_idx - 3):start_idx]
    if len(meta_candidates) >= 1:
        title = row_signature(meta_candidates[0])[-1] if row_signature(meta_candidates[0]) else None
    if len(meta_candidates) >= 2:
        desc_values = row_signature(meta_candidates[1])
        description = desc_values[-1] if len(desc_values) > 1 else None
    if len(meta_candidates) >= 3:
        prereq_values = row_signature(meta_candidates[2])
        prerequisites = prereq_values[-1] if len(prereq_values) > 1 else None

    level_headers: list[tuple[int, str]] = []
    meta_headers: list[tuple[int, str]] = []
    for index, cell in enumerate(level_row):
        value = normalize_text(cell)
        if not value:
            continue
        if is_level_header(value):
            level_headers.append((index, value))
        else:
            meta_headers.append((index, value))

    if not level_headers and previous_level_headers:
        level_headers = previous_level_headers[:]
        issues.append("level_headers_inherited_from_previous_block")

    if not title:
        issues.append("missing_block_title")
    if not level_headers:
        issues.append("missing_level_headers")

    header = BlockHeader(
        header_row_idx=start_idx,
        level_row_idx=min(start_idx + 1, len(rows) - 1),
        title=title,
        description=description,
        prerequisites=prerequisites,
        skill_col=skill_col,
        dimension_col=dimension_col,
        number_col=number_col,
        level_headers=level_headers,
        meta_headers=meta_headers,
    )
    return header, issues


class CatalogBuilder:
    def __init__(self, conn: sqlite3.Connection, schema_path: Path):
        self.conn = conn
        self.conn.row_factory = sqlite3.Row
        self.schema_path = schema_path
        self.dimension_cache: dict[str, int] = {}
        self.profile_cache: dict[str, int] = {}
        self.scale_cache: dict[str, int] = {}
        self.level_cache: dict[tuple[int, str], int] = {}
        self.skill_cache: dict[str, int] = {}
        self.skill_alias_cache: set[tuple[int, str]] = set()
        self.competency_cache: dict[str, int] = {}

    def initialize(self) -> None:
        self.conn.executescript(self.schema_path.read_text(encoding="utf-8"))
        for code, title in {
            "knowledge": "Знает",
            "ability": "Умеет",
            "proficiency": "Владеет",
            "understanding": "Понимает",
            "unspecified": "Не указано",
        }.items():
            self.conn.execute(
                "INSERT OR IGNORE INTO dimension(code, title) VALUES(?, ?)",
                (code, title),
            )
        self.conn.commit()
        for row in self.conn.execute("SELECT id, code FROM dimension"):
            self.dimension_cache[row["code"]] = row["id"]

    def create_ingest_run(self, source_root: Path) -> int:
        cursor = self.conn.execute(
            "INSERT INTO ingest_run(source_root, status) VALUES(?, 'running')",
            (str(source_root),),
        )
        return int(cursor.lastrowid)

    def finish_ingest_run(self, run_id: int, summary: dict[str, object], status: str = "completed") -> None:
        self.conn.execute(
            """
            UPDATE ingest_run
            SET finished_at = ?, status = ?, summary_json = ?
            WHERE id = ?
            """,
            (
                datetime.now(timezone.utc).isoformat(),
                status,
                json.dumps(summary, ensure_ascii=False, indent=2),
                run_id,
            ),
        )
        self.conn.commit()

    def add_review(
        self,
        entity_type: str,
        entity_id: int | None,
        source_ref: str,
        reason_code: str,
        severity: str,
        details: str,
    ) -> None:
        self.conn.execute(
            """
            INSERT INTO review_queue(entity_type, entity_id, source_ref, reason_code, severity, details)
            VALUES(?, ?, ?, ?, ?, ?)
            """,
            (
                entity_type,
                entity_id,
                source_ref,
                reason_code,
                severity,
                humanize_review_details(reason_code, source_ref, details),
            ),
        )

    def get_or_create_profile(self, name: str, source_kind: str) -> int:
        slug = slugify(name)
        cache_key = f"{source_kind}:{slug}"
        if cache_key in self.profile_cache:
            return self.profile_cache[cache_key]
        cursor = self.conn.execute(
            """
            INSERT INTO profile(slug, name, source_kind)
            VALUES(?, ?, ?)
            ON CONFLICT(slug) DO UPDATE SET name = excluded.name
            RETURNING id
            """,
            (slug, name, source_kind),
        )
        profile_id = int(cursor.fetchone()["id"])
        self.profile_cache[cache_key] = profile_id
        return profile_id

    def get_or_create_competency(self, title: str | None, description: str | None) -> int:
        normalized = normalize_key(title) if title else ""
        normalized = normalized or "__missing__"
        if normalized in self.competency_cache:
            return self.competency_cache[normalized]
        cursor = self.conn.execute(
            """
            INSERT INTO competency(normalized_title, title, description, status)
            VALUES(?, ?, ?, ?)
            ON CONFLICT(normalized_title) DO UPDATE SET
                title = COALESCE(competency.title, excluded.title),
                description = COALESCE(competency.description, excluded.description)
            RETURNING id
            """,
            (
                normalized,
                title or "[MISSING TITLE]",
                description,
                "candidate" if title is None else "active",
            ),
        )
        competency_id = int(cursor.fetchone()["id"])
        self.competency_cache[normalized] = competency_id
        return competency_id

    def get_or_create_skill(self, raw_name: str) -> int:
        normalized = normalize_key(raw_name)
        if normalized in self.skill_cache:
            return self.skill_cache[normalized]
        cursor = self.conn.execute(
            """
            INSERT INTO skill(normalized_name, canonical_name, skill_type, status)
            VALUES(?, ?, 'unknown', 'active')
            ON CONFLICT(normalized_name) DO UPDATE SET canonical_name = skill.canonical_name
            RETURNING id
            """,
            (normalized, raw_name.strip()),
        )
        skill_id = int(cursor.fetchone()["id"])
        self.skill_cache[normalized] = skill_id
        self.add_skill_alias(skill_id, raw_name, "source")
        return skill_id

    def add_skill_alias(self, skill_id: int, alias: str, source: str) -> None:
        normalized_alias = normalize_key(alias)
        cache_key = (skill_id, normalized_alias)
        if cache_key in self.skill_alias_cache:
            return
        self.conn.execute(
            """
            INSERT OR IGNORE INTO skill_alias(skill_id, alias, normalized_alias, source)
            VALUES(?, ?, ?, ?)
            """,
            (skill_id, alias.strip(), normalized_alias, source),
        )
        self.skill_alias_cache.add(cache_key)

    def get_or_create_scale(self, level_headers: list[tuple[int, str]]) -> tuple[int | None, dict[int, int]]:
        if not level_headers:
            return None, {}
        signature = " || ".join(normalize_key(label) for _, label in level_headers)
        if signature not in self.scale_cache:
            scale_code = f"scale-{hashlib.sha1(signature.encode('utf-8')).hexdigest()[:10]}"
            scale_title = " / ".join(label for _, label in level_headers)
            cursor = self.conn.execute(
                """
                INSERT INTO proficiency_scale(code, title, normalized_signature)
                VALUES(?, ?, ?)
                ON CONFLICT(normalized_signature) DO UPDATE SET title = excluded.title
                RETURNING id
                """,
                (scale_code, scale_title, signature),
            )
            scale_id = int(cursor.fetchone()["id"])
            self.scale_cache[signature] = scale_id
        scale_id = self.scale_cache[signature]

        mapping: dict[int, int] = {}
        for order, (column_index, label) in enumerate(level_headers, start=1):
            cache_key = (scale_id, normalize_key(label))
            if cache_key not in self.level_cache:
                cursor = self.conn.execute(
                    """
                    INSERT INTO proficiency_level(scale_id, code, title, sort_order, canonical_band)
                    VALUES(?, ?, ?, ?, ?)
                    ON CONFLICT(scale_id, title) DO UPDATE SET
                        sort_order = excluded.sort_order,
                        canonical_band = COALESCE(proficiency_level.canonical_band, excluded.canonical_band)
                    RETURNING id
                    """,
                    (scale_id, slugify(label), label, order, canonical_band(label)),
                )
                self.level_cache[cache_key] = int(cursor.fetchone()["id"])
            mapping[column_index] = self.level_cache[cache_key]
        return scale_id, mapping

    def dimension_id(self, raw_dimension: str | None) -> int:
        if raw_dimension:
            normalized = normalize_key(raw_dimension)
            if normalized in DIMENSION_MAP:
                return self.dimension_cache[DIMENSION_MAP[normalized][0]]
            slug = slugify(raw_dimension)
            if slug not in self.dimension_cache:
                self.conn.execute(
                    "INSERT OR IGNORE INTO dimension(code, title) VALUES(?, ?)",
                    (slug, raw_dimension),
                )
                row = self.conn.execute("SELECT id FROM dimension WHERE code = ?", (slug,)).fetchone()
                self.dimension_cache[slug] = int(row["id"])
            return self.dimension_cache[slug]
        return self.dimension_cache["unspecified"]

    def source_ref(self, workbook: str, sheet: str, row_no: int | None = None) -> str:
        if row_no is None:
            return f"{workbook}::{sheet}"
        return f"{workbook}::{sheet}::row-{row_no}"


def insert_workbook(
    builder: CatalogBuilder,
    run_id: int,
    workbook_path: Path,
) -> tuple[int, int]:
    source_kind = detect_source_kind(workbook_path.name)
    modified = datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=timezone.utc).isoformat()
    cursor = builder.conn.execute(
        """
        INSERT INTO source_workbook(ingest_run_id, file_path, file_name, sha256, last_modified_utc, source_kind)
        VALUES(?, ?, ?, ?, ?, ?)
        """,
        (run_id, str(workbook_path), workbook_path.name, sha256_file(workbook_path), modified, source_kind),
    )
    workbook_id = int(cursor.lastrowid)

    profile_name = clean_profile_name(workbook_path.name)
    profile_id = builder.get_or_create_profile(profile_name, source_kind)
    builder.conn.execute(
        """
        INSERT OR IGNORE INTO profile_source(profile_id, source_workbook_id, version_label, is_primary)
        VALUES(?, ?, ?, 1)
        """,
        (profile_id, workbook_id, Path(workbook_path.name).stem),
    )
    return workbook_id, profile_id


def parse_workbook(
    builder: CatalogBuilder,
    workbook_path: Path,
    workbook_id: int,
    profile_id: int,
) -> None:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    for sheet_order, sheet_name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[sheet_name]
        skipped = is_skip_sheet(sheet_name)
        cursor = builder.conn.execute(
            """
            INSERT INTO source_sheet(source_workbook_id, sheet_name, sheet_order, is_skipped, skip_reason)
            VALUES(?, ?, ?, ?, ?)
            """,
            (
                workbook_id,
                sheet_name,
                sheet_order,
                1 if skipped else 0,
                "auxiliary_sheet" if skipped else None,
            ),
        )
        sheet_id = int(cursor.lastrowid)
        if skipped:
            continue

        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        block_starts = find_block_starts(rows)
        previous_level_headers: list[tuple[int, str]] | None = None
        if not block_starts:
            builder.add_review(
                "sheet",
                sheet_id,
                builder.source_ref(workbook_path.name, sheet_name),
                "no_header_rows",
                "warning",
                "На листе не найдено ни одного блока с заголовком Skills/Х.",
            )
            continue

        for block_no, start_idx in enumerate(block_starts, start=1):
            end_idx = block_starts[block_no] if block_no < len(block_starts) else len(rows)
            block_header, header_issues = build_block_header(rows, start_idx, previous_level_headers)
            if block_header.level_headers:
                previous_level_headers = block_header.level_headers

            scale_id, level_mapping = builder.get_or_create_scale(block_header.level_headers)

            block_cursor = builder.conn.execute(
                """
                INSERT INTO source_block(
                    source_sheet_id,
                    block_no,
                    header_row_number,
                    level_row_number,
                    end_row_number,
                    raw_title,
                    raw_description,
                    raw_prerequisites,
                    raw_scale_signature
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sheet_id,
                    block_no,
                    block_header.header_row_idx + 1,
                    block_header.level_row_idx + 1,
                    end_idx,
                    block_header.title,
                    block_header.description,
                    block_header.prerequisites,
                    " || ".join(label for _, label in block_header.level_headers) if block_header.level_headers else None,
                ),
            )
            source_block_id = int(block_cursor.lastrowid)

            competency_id = builder.get_or_create_competency(block_header.title, block_header.description)
            pc_cursor = builder.conn.execute(
                """
                INSERT INTO profile_competency(
                    profile_id,
                    competency_id,
                    source_block_id,
                    scale_id,
                    title_in_source,
                    description_in_source,
                    prerequisites_text,
                    sort_order,
                    review_state
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    profile_id,
                    competency_id,
                    source_block_id,
                    scale_id,
                    block_header.title,
                    block_header.description,
                    block_header.prerequisites,
                    block_no,
                    "needs_review" if header_issues else "accepted",
                ),
            )
            profile_competency_id = int(pc_cursor.lastrowid)

            for issue in header_issues:
                builder.add_review(
                    "block",
                    source_block_id,
                    builder.source_ref(workbook_path.name, sheet_name, block_header.header_row_idx + 1),
                    issue,
                    "warning",
                    f"Проблема структуры блока '{block_header.title or '[без названия]'}'.",
                )

            if block_header.title and "?" in block_header.title:
                builder.add_review(
                    "block",
                    source_block_id,
                    builder.source_ref(workbook_path.name, sheet_name, block_header.header_row_idx + 1),
                    "ambiguous_block_title",
                    "warning",
                    f"Подозрительное название блока: {block_header.title}",
                )

            current_skill_name: str | None = None
            current_dimension: str | None = None
            current_competency_skill_id: int | None = None
            skill_order = 0

            for row_index in range(start_idx + 2, end_idx):
                row = rows[row_index]
                if not row_has_any_content(row):
                    continue

                raw_skill = normalize_text(row[block_header.skill_col]) if block_header.skill_col < len(row) else None
                raw_dimension = normalize_text(row[block_header.dimension_col]) if block_header.dimension_col < len(row) else None
                raw_number = normalize_text(row[block_header.number_col]) if block_header.number_col < len(row) else None

                inherited_skill = 0
                inherited_dimension = 0

                if raw_skill:
                    current_skill_name = raw_skill
                    skill_order += 1
                    skill_id = builder.get_or_create_skill(raw_skill)
                    current_competency_skill_id = int(
                        builder.conn.execute(
                            """
                            INSERT INTO competency_skill(
                                profile_competency_id,
                                skill_id,
                                source_skill_name,
                                skill_order,
                                review_state
                            )
                            VALUES(?, ?, ?, ?, ?)
                            """,
                            (
                                profile_competency_id,
                                skill_id,
                                raw_skill.strip(),
                                skill_order,
                                "needs_review" if "?" in raw_skill else "accepted",
                            ),
                        ).lastrowid
                    )
                    if raw_skill != raw_skill.strip():
                        builder.add_review(
                            "skill",
                            skill_id,
                            builder.source_ref(workbook_path.name, sheet_name, row_index + 1),
                            "skill_name_trimmed",
                            "info",
                            f"Лишние пробелы вокруг названия навыка: '{raw_skill}'.",
                        )
                    if "?" in raw_skill:
                        builder.add_review(
                            "skill",
                            skill_id,
                            builder.source_ref(workbook_path.name, sheet_name, row_index + 1),
                            "ambiguous_skill_name",
                            "warning",
                            f"Неоднозначное название навыка: '{raw_skill}'.",
                        )
                else:
                    inherited_skill = 1

                if raw_dimension:
                    current_dimension = raw_dimension
                else:
                    inherited_dimension = 1

                base_text = None
                meta_values: list[tuple[str, str]] = []
                for column_index, meta_label in block_header.meta_headers:
                    if column_index >= len(row):
                        continue
                    value = normalize_text(row[column_index])
                    if value is None:
                        continue
                    if normalize_key(meta_label) == "описание навыка":
                        base_text = value
                    else:
                        meta_values.append((meta_label, value))

                level_values: list[tuple[int, str, str]] = []
                for column_index, level_label in block_header.level_headers:
                    if column_index >= len(row):
                        continue
                    value = normalize_text(row[column_index])
                    if value is None:
                        continue
                    level_values.append((column_index, level_label, value))

                has_payload = bool(base_text or meta_values or level_values)
                if not any((raw_skill, raw_dimension, base_text, meta_values, level_values)):
                    continue

                if not has_payload:
                    continue

                if current_competency_skill_id is None or current_skill_name is None:
                    builder.add_review(
                        "indicator_row",
                        None,
                        builder.source_ref(workbook_path.name, sheet_name, row_index + 1),
                        "orphan_indicator_row",
                        "error",
                        "Строка содержит индикаторы, но не удалось определить текущий навык.",
                    )
                    continue

                dimension_id = builder.dimension_id(current_dimension)
                indicator_cursor = builder.conn.execute(
                    """
                    INSERT INTO indicator_row(
                        competency_skill_id,
                        dimension_id,
                        source_row_number,
                        inherited_skill,
                        inherited_dimension,
                        base_text,
                        raw_number
                    )
                    VALUES(?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        current_competency_skill_id,
                        dimension_id,
                        row_index + 1,
                        inherited_skill,
                        inherited_dimension,
                        base_text,
                        raw_number,
                    ),
                )
                indicator_row_id = int(indicator_cursor.lastrowid)

                if inherited_dimension and current_dimension is None:
                    builder.add_review(
                        "indicator_row",
                        indicator_row_id,
                        builder.source_ref(workbook_path.name, sheet_name, row_index + 1),
                        "missing_dimension",
                        "warning",
                        "У строки нет явного измерения (знает/умеет/владеет) и его нельзя унаследовать.",
                    )

                for meta_label, meta_value in meta_values:
                    builder.conn.execute(
                        """
                        INSERT INTO indicator_row_meta(indicator_row_id, meta_key, meta_value)
                        VALUES(?, ?, ?)
                        """,
                        (indicator_row_id, meta_label, meta_value),
                    )

                if not level_values and base_text:
                    builder.add_review(
                        "indicator_row",
                        indicator_row_id,
                        builder.source_ref(workbook_path.name, sheet_name, row_index + 1),
                        "base_text_without_levels",
                        "info",
                        "Есть базовый текст индикатора, но нет явных значений по уровням.",
                    )

                for sort_order, (column_index, level_label, raw_value) in enumerate(level_values, start=1):
                    lowered = normalize_key(raw_value)
                    if lowered in {"+", '"+"'}:
                        kind = "marker_plus"
                    elif lowered in {"-", '"-"'}:
                        kind = "marker_minus"
                    elif re.fullmatch(r"-?\d+(?:[.,]\d+)?", lowered):
                        kind = "numeric"
                    else:
                        kind = "text"
                    builder.conn.execute(
                        """
                        INSERT INTO indicator_level_cell(
                            indicator_row_id,
                            proficiency_level_id,
                            raw_level_label,
                            raw_value,
                            value_kind,
                            sort_order
                        )
                        VALUES(?, ?, ?, ?, ?, ?)
                        """,
                        (
                            indicator_row_id,
                            level_mapping.get(column_index),
                            level_label,
                            raw_value,
                            kind,
                            sort_order,
                        ),
                    )


def collect_summary(conn: sqlite3.Connection) -> dict[str, object]:
    def scalar(query: str) -> int:
        return int(conn.execute(query).fetchone()[0])

    duplicate_skills = [
        dict(row)
        for row in conn.execute(
            """
            SELECT canonical_name, profile_count, competency_links
            FROM v_skill_usage
            ORDER BY profile_count DESC, competency_links DESC, canonical_name
            LIMIT 15
            """
        )
    ]

    review_breakdown = [
        dict(row)
        for row in conn.execute(
            """
            SELECT reason_code, severity, COUNT(*) AS cnt
            FROM review_queue
            GROUP BY reason_code, severity
            ORDER BY cnt DESC, reason_code
            """
        )
    ]

    scales = [
        dict(row)
        for row in conn.execute(
            """
            SELECT ps.title, COUNT(pl.id) AS levels
            FROM proficiency_scale ps
            LEFT JOIN proficiency_level pl ON pl.scale_id = ps.id
            GROUP BY ps.id, ps.title
            ORDER BY ps.title
            """
        )
    ]

    return {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "counts": {
            "workbooks": scalar("SELECT COUNT(*) FROM source_workbook"),
            "sheets": scalar("SELECT COUNT(*) FROM source_sheet"),
            "blocks": scalar("SELECT COUNT(*) FROM source_block"),
            "profiles": scalar("SELECT COUNT(*) FROM profile"),
            "competencies": scalar("SELECT COUNT(*) FROM competency"),
            "profile_competencies": scalar("SELECT COUNT(*) FROM profile_competency"),
            "skills": scalar("SELECT COUNT(*) FROM skill"),
            "competency_skills": scalar("SELECT COUNT(*) FROM competency_skill"),
            "indicator_rows": scalar("SELECT COUNT(*) FROM indicator_row"),
            "indicator_level_cells": scalar("SELECT COUNT(*) FROM indicator_level_cell"),
            "open_reviews": scalar("SELECT COUNT(*) FROM v_pending_reviews"),
        },
        "duplicate_skills": duplicate_skills,
        "review_breakdown": review_breakdown,
        "scales": scales,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a local skills catalog SQLite database from Excel competency profiles.")
    parser.add_argument(
        "--input-root",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "КПшки" / "КПшки",
        help="Directory with source .xlsx competency profiles.",
    )
    parser.add_argument(
        "--output-db",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "artifacts" / "skills_catalog.sqlite",
        help="Path to the output SQLite database.",
    )
    parser.add_argument(
        "--summary-json",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "artifacts" / "catalog_summary.json",
        help="Path to the generated audit summary JSON.",
    )
    args = parser.parse_args()

    input_root = args.input_root.resolve()
    output_db = args.output_db.resolve()
    summary_json = args.summary_json.resolve()
    schema_path = Path(__file__).resolve().parents[1] / "sql" / "catalog_schema.sql"

    output_db.parent.mkdir(parents=True, exist_ok=True)
    summary_json.parent.mkdir(parents=True, exist_ok=True)
    if output_db.exists():
        output_db.unlink()

    conn = sqlite3.connect(output_db)
    try:
        builder = CatalogBuilder(conn, schema_path)
        builder.initialize()
        run_id = builder.create_ingest_run(input_root)

        for workbook_path in workbook_paths(input_root):
            workbook_id, profile_id = insert_workbook(builder, run_id, workbook_path)
            parse_workbook(builder, workbook_path, workbook_id, profile_id)

        conn.commit()
        apply_methodologist_review_decisions(conn)
        summary = collect_summary(conn)
        builder.finish_ingest_run(run_id, summary, status="completed")
        summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        conn.rollback()
        if "builder" in locals() and "run_id" in locals():
            builder.finish_ingest_run(run_id, {"error": str(exc)}, status="failed")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
