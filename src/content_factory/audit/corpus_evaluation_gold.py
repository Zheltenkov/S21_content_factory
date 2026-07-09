"""Gold corpus loading and normalization for audit evaluation."""

from __future__ import annotations

import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from content_factory.audit.corpus_evaluation_models import GoldCorpusCase, GoldCorpusItem, _ProjectCandidate
from content_factory.audit.domain import CRITERION_LABELS, Criterion

PROJECT_COLUMN = "Проект"
PROBLEM_COLUMN = "Проблема"
DETAILS_COLUMN = "Детали"

CRITERION_ALIASES: dict[Criterion, tuple[str, ...]] = {
    Criterion.LINKS: (
        "ссылка",
        "url",
        "api",
        "сервис",
    ),
    Criterion.TECHNOLOGY_FRESHNESS: (
        "устар",
        "версии ос",
        "инструмент",
        "библиотек",
        "технолог",
        "фреймворк",
        "стек",
    ),
    Criterion.FACTS: (
        "факт",
        "определени",
        "пример",
        "логик",
        "противореч",
        "неверн",
    ),
    Criterion.READABILITY: (
        "опечат",
        "граммат",
        "орфограф",
        "формулиров",
        "нумерац",
        "двоеточ",
    ),
    Criterion.CHECKLIST_ALIGNMENT: (
        "чек-лист",
        "чеклист",
        "check-list",
        "несоответствие задания чек-листу",
        "несоответствие задания чеклисту",
    ),
    Criterion.CORRECTNESS: (
        "ошибка в задании",
        "нарушение логики",
        "логик",
        "противореч",
        "некоррект",
        "неверн",
        "ошибка в доп материалах",
        "ошибка в дополнительных материалах",
    ),
}


def load_gold_items(
    gold_xlsx_path: Path,
    unit_candidates: list[_ProjectCandidate],
) -> tuple[list[GoldCorpusItem], list[str]]:
    """Читает Excel и переводит строки в эталонные критерии."""

    workbook = load_workbook(gold_xlsx_path, data_only=True)
    sheet = workbook.active
    header = _header_map(sheet)
    required_columns = {PROJECT_COLUMN, PROBLEM_COLUMN, DETAILS_COLUMN}
    missing_columns = sorted(required_columns - set(header))
    if missing_columns:
        raise ValueError(f"В Excel не найдены обязательные колонки: {', '.join(missing_columns)}")

    items: list[GoldCorpusItem] = []
    notes: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_project = _cell_text(row, header[PROJECT_COLUMN])
        raw_problem = _cell_text(row, header[PROBLEM_COLUMN])
        details = _cell_text(row, header[DETAILS_COLUMN])
        if not raw_project and not raw_problem and not details:
            continue

        candidate, score = _match_project(raw_project, unit_candidates)
        criteria = _criteria_from_gold_row(raw_problem, details)
        if not criteria:
            notes.append(f"Строка {row_number}: не удалось вывести критерий из типа проблемы {raw_problem!r}.")
            continue
        if score < 0.55:
            notes.append(
                f"Строка {row_number}: слабое сопоставление проекта {raw_project!r} "
                f"с папкой {candidate.raw_name!r}, score={score:.2f}."
            )

        items.append(
            GoldCorpusItem(
                row_number=row_number,
                raw_project=raw_project,
                matched_project=candidate.raw_name,
                project_id=candidate.project_id,
                raw_problem=raw_problem,
                details=details,
                criteria=criteria,
            )
        )
    return items, notes


def gold_cases_from_items(items: list[GoldCorpusItem]) -> list[GoldCorpusCase]:
    """Разбивает Excel-строки на атомарные эталонные случаи."""

    cases: list[GoldCorpusCase] = []
    for item in items:
        detail_cases = _split_gold_detail_cases(item.details or item.raw_problem)
        for index, detail in enumerate(detail_cases, start=1):
            detail_criteria = _criteria_from_gold_case(str(detail["text"])) or item.criteria
            for criterion in detail_criteria:
                cases.append(
                    GoldCorpusCase(
                        case_id=f"gold_{item.row_number}_{criterion}_{index}",
                        row_number=item.row_number,
                        raw_project=item.raw_project,
                        matched_project=item.matched_project,
                        project_id=item.project_id,
                        criterion=criterion,
                        line_start=detail["line_start"],
                        line_end=detail["line_end"],
                        gold_text=detail["text"],
                    )
                )
    return cases


def normalize_project_name(value: str) -> str:
    """Нормализует имя проекта для сопоставления Excel и папок."""

    text = value.lower().replace("с", "c")
    text = re.sub(r"\.id_\d+.*$", "", text)
    text = re.sub(r"\(\d+\)", "", text)
    text = re.sub(r"[^a-zа-яё0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def project_tokens(value: str) -> list[str]:
    """Выделяет значимые токены имени проекта."""

    normalized = normalize_project_name(value)
    return [token for token in normalized.split() if token not in {"master", "id"}]


def _criteria_from_gold_case(text: str) -> list[str]:
    """Выводит критерий из конкретной строки эталонной ошибки, без размножения на весь тип проблемы."""

    lowered = str(text or "").lower()
    markers: tuple[tuple[Criterion, tuple[str, ...]], ...] = (
        (
            Criterion.LINKS,
            (
                "сломанная ссылка",
                "ссылка",
                "url",
                "http",
                "https",
            ),
        ),
        (
            Criterion.TECHNOLOGY_FRESHNESS,
            (
                "устар",
                "неактуаль",
                "версии ос",
                "библиотек",
                "фреймворк",
                "стек",
            ),
        ),
        (
            Criterion.CHECKLIST_ALIGNMENT,
            (
                "чеклист",
                "чек-лист",
                "чек лист",
                "check-list",
                "checklist",
            ),
        ),
        (
            Criterion.READABILITY,
            (
                "опечат",
                "двоеточ",
                "нумерац",
                "пронумер",
                "граммат",
                "тавтолог",
                "формулиров",
                "кавыч",
                "input ",
                "output",
                "example",
                "result",
            ),
        ),
        (
            Criterion.CORRECTNESS,
            (
                "противореч",
                "некоррект",
                "неверн",
                "не является",
                "по факту",
                "ошибка в задании",
            ),
        ),
    )
    result: list[Criterion] = []
    for criterion, aliases in markers:
        if any(alias in lowered for alias in aliases):
            result.append(criterion)
    return [criterion.value for criterion in dict.fromkeys(result)]


def _split_gold_detail_cases(text: str) -> list[dict[str, Any]]:
    """Выделяет отдельные ошибки из многострочного описания разметки."""

    lines = _normalise_gold_detail_lines(text)
    if not lines:
        lines = [str(text or "").strip()]
    cases: list[dict[str, Any]] = []
    for line in lines:
        line_start, line_end = _line_range_from_text(line)
        cases.append(
            {
                "line_start": line_start,
                "line_end": line_end,
                "text": _strip_gold_list_marker(line),
            }
        )
    if cases:
        return cases
    line_start, line_end = _line_range_from_text(text)
    return [{"line_start": line_start, "line_end": line_end, "text": str(text or "").strip()}]


def _normalise_gold_detail_lines(text: str) -> list[str]:
    """Оставляет только строки с дефектами, отделяя решения, доказательства и шапки таблиц."""

    result: list[str] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip().strip('"').strip()
        if not line or _looks_like_gold_metadata_line(line):
            continue
        if _starts_non_defect_gold_section(line):
            break
        result.append(line)
    return result


def _looks_like_gold_metadata_line(value: str) -> bool:
    """Отсекает служебные строки из Excel-описания, если они попали в детали."""

    lowered = re.sub(r"\s+", " ", value.lower().strip(" :"))
    exact_headings = {
        "строка - проблема - решение",
        "строка проблема решение",
        "проблема - решение",
        "problem - solution",
        "line - problem - solution",
    }
    if lowered in exact_headings:
        return True
    return lowered.startswith(("тип проблемы", "проект", "критерий", "архив", ":small_blue_diamond: description"))


def _starts_non_defect_gold_section(value: str) -> bool:
    """Определяет начало раздела с решением или доказательствами, а не с эталонной ошибкой."""

    lowered = re.sub(r"\s+", " ", value.lower().strip(" :"))
    prefixes = (
        "предложение по решению",
        "предлагаемое решение",
        "решение",
        "аргументы и исследования",
        "аргументы",
        "исследования",
        "доказательства",
        "подтверждение",
        "скрины",
        "скриншоты",
    )
    return lowered.startswith(prefixes)


def _strip_gold_list_marker(value: str) -> str:
    """Убирает маркер списка, не путая его с номером строки исходного файла."""

    text = str(value or "").strip()
    bullet_match = re.match(r"^\s*[*•]\s*(.+)$", text)
    if bullet_match is not None:
        return bullet_match.group(1).strip()
    ordered_match = re.match(r"^\s*\d{1,3}[.)]\s+(.+)$", text)
    if ordered_match is not None:
        return ordered_match.group(1).strip()
    return text


def _line_range_from_text(value: str) -> tuple[int | None, int | None]:
    """Достаёт строку или диапазон строк из эталонного описания."""

    text = str(value or "")
    patterns = (
        r"^\s*(?:строк[аеи]\s*)?(\d{1,5})\s*[–—-]\s*(\d{1,5})\s+[–—-]",
        r"^\s*(?:строк[аеи]\s*)?(\d{1,5})\s+[–—-]",
        r"\bстрок[аеи]?\s*(\d{1,5})\s*[–—-]\s*(\d{1,5})\b",
        r"\bстрок[аеи]?\s*(\d{1,5})\b(?=\s*[:.)]|\s|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match is None:
            continue
        start = int(match.group(1))
        end = int(match.group(2)) if match.lastindex and match.lastindex >= 2 and match.group(2) else start
        return min(start, end), max(start, end)
    return None, None


def _header_map(sheet: Any) -> dict[str, int]:
    """Возвращает индексы колонок по первой строке Excel."""

    result: dict[str, int] = {}
    for index, cell in enumerate(sheet[1]):
        if cell.value:
            result[str(cell.value).strip()] = index
    return result


def _cell_text(row: tuple[Any, ...], index: int) -> str:
    """Безопасно достаёт текст ячейки."""

    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _match_project(raw_project: str, candidates: list[_ProjectCandidate]) -> tuple[_ProjectCandidate, float]:
    """Сопоставляет короткое имя проекта из Excel с фактической папкой."""

    normalized = normalize_project_name(raw_project)
    tokens = set(project_tokens(raw_project))
    best: tuple[_ProjectCandidate, float] | None = None
    for candidate in candidates:
        score = _project_match_score(normalized, tokens, candidate)
        if best is None or score > best[1]:
            best = (candidate, score)
    if best is None:
        raise ValueError("В отчёте нет единиц контента для сопоставления с Excel.")
    return best


def _project_match_score(normalized: str, tokens: set[str], candidate: _ProjectCandidate) -> float:
    """Считает устойчивый score для разных написаний имени проекта."""

    if not normalized:
        return 0.0
    if normalized == candidate.normalized_name:
        return 1.0
    if normalized in candidate.normalized_name or candidate.normalized_name in normalized:
        return 0.95
    token_overlap = len(tokens & set(candidate.tokens)) / max(len(tokens), 1)
    sequence_score = SequenceMatcher(a=normalized, b=candidate.normalized_name).ratio()
    prefix_score = 0.0
    if tokens and candidate.tokens and next(iter(tokens)) in candidate.tokens:
        prefix_score = 0.25
    return max(sequence_score, token_overlap + prefix_score)


def _criteria_from_gold_row(raw_problem: str, details: str) -> list[str]:
    """Выводит наши критерии из свободного описания проблемы."""

    # The gold sheet already states the criterion in the "Проблема" column;
    # trust an exact label match before falling back to fuzzy alias matching.
    label_to_criterion = {label.lower(): criterion for criterion, label in CRITERION_LABELS.items()}
    direct = label_to_criterion.get(raw_problem.strip().lower())
    if direct is not None and direct != Criterion.ACTUALITY:
        return [direct.value]

    problem_text = raw_problem.lower()
    detail_text = details.lower()
    criteria: list[Criterion] = []

    # Тип проблемы надёжнее деталей, поэтому сначала используем его.
    for criterion, aliases in CRITERION_ALIASES.items():
        if any(alias in problem_text for alias in aliases):
            criteria.append(criterion)

    # Детали добавляют критерии только по сильным маркерам, чтобы не раздувать эталон.
    detail_markers: dict[Criterion, tuple[str, ...]] = {
        Criterion.LINKS: ("сломанная ссылка", "битая ссылка", "неверная ссылка", "url", "http", "https"),
        Criterion.TECHNOLOGY_FRESHNESS: ("неактуаль", "устар", "старый стандарт", "версии ос", "библиотек", "стек"),
        Criterion.FACTS: ("факт", "определение", "пример", "таблица вывода", "расхождение"),
        Criterion.CHECKLIST_ALIGNMENT: ("в чеклисте", "в чек-листе", "чеклист", "чек лист"),
        Criterion.READABILITY: ("опечат", "нумерация", "пронумер", "грамматика"),
        Criterion.CORRECTNESS: ("противореч", "некоррект", "по факту", "не является", "отсутствует"),
    }
    for criterion, markers in detail_markers.items():
        if any(marker in detail_text for marker in markers):
            criteria.append(criterion)

    return [criterion.value for criterion in dict.fromkeys(criteria)]
