"""Оценка качества аудита на корпусе проектов с Excel-разметкой."""

from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from content_factory.audit.domain import AuditReport, CRITERION_LABELS, Criterion, Finding


PROJECT_COLUMN = "Проект"
PROBLEM_COLUMN = "Проблема"
DETAILS_COLUMN = "Детали"


CRITERION_ALIASES: dict[Criterion, tuple[str, ...]] = {
    Criterion.LINKS: (
        "ссылка",
        "url",
        "api",
        "сервис",
    ),
    Criterion.TECHNOLOGY_FRESHNESS: (
        "устар",
        "версии ос",
        "инструмент",
        "библиотек",
        "технолог",
        "фреймворк",
        "стек",
    ),
    Criterion.FACTS: (
        "факт",
        "определени",
        "пример",
        "логик",
        "противореч",
        "неверн",
    ),
    Criterion.READABILITY: (
        "опечат",
        "граммат",
        "орфограф",
        "формулиров",
        "нумерац",
        "двоеточ",
    ),
    Criterion.CHECKLIST_ALIGNMENT: (
        "чек-лист",
        "чеклист",
        "check-list",
        "несоответствие задания чек-листу",
        "несоответствие задания чеклисту",
    ),
    Criterion.CORRECTNESS: (
        "ошибка в задании",
        "нарушение логики",
        "логик",
        "противореч",
        "некоррект",
        "неверн",
        "ошибка в доп материалах",
        "ошибка в дополнительных материалах",
    ),
}


CHECKER_GROUPS: dict[str, str] = {
    "broken_url_syntax_checker": "deterministic_rules",
    "label_punctuation_checker": "deterministic_rules",
    "local_consistency_checker": "deterministic_rules",
    "markdown_structure_checker": "deterministic_rules",
    "spelling_wording_checker": "editorial_rules",
    "link_checker": "links_and_resources",
    "local_link_checker": "links_and_resources",
    "resource_availability_checker": "links_and_resources",
    "checklist_checker": "checklist_and_artifacts",
    "fact_checker_perplexity": "factcheck",
    "readme_fact_actuality_checker": "factcheck",
    "tech_freshness_checker": "factcheck",
    "dependency_freshness_checker": "factcheck",
    "curriculum_relevance_checker": "methodology",
    "market_fit_checker": "methodology",
    "model_rubric_checker": "methodology",
}

CHECKER_GROUP_LABELS: dict[str, str] = {
    "deterministic_rules": "Детерминированные правила",
    "editorial_rules": "Редакторские правила",
    "links_and_resources": "Ссылки и ресурсы",
    "checklist_and_artifacts": "Чек-лист и артефакты",
    "factcheck": "Фактчек и актуальность",
    "methodology": "Методические критерии",
    "other": "Прочие проверки",
}


class CorpusEvaluationKey(BaseModel, frozen=True):
    """Ключ сравнения: один проект и один критерий."""

    project_id: str
    criterion: str


class GoldCorpusItem(BaseModel):
    """Одна эталонная строка после нормализации Excel."""

    row_number: int
    raw_project: str
    matched_project: str
    project_id: str
    raw_problem: str
    details: str
    criteria: list[str]


class GoldCorpusCase(BaseModel):
    """Один атомарный эталонный случай: проект, критерий и конкретная строка/описание."""

    case_id: str
    row_number: int
    raw_project: str
    matched_project: str
    project_id: str
    criterion: str
    line_start: int | None = None
    line_end: int | None = None
    gold_text: str
    file_hint: str | None = None


class PredictedCorpusItem(BaseModel):
    """Один найденный алгоритмом случай в формате, удобном для сопоставления."""

    finding_id: str
    project_id: str
    project: str
    criterion: str
    checker_name: str
    line_start: int | None = None
    line_end: int | None = None
    file_path: str | None = None
    severity: str | None = None
    verdict: str | None = None
    confidence: float | None = None
    issue_type: str | None = None
    found_text: str


class CorpusEvaluationMatch(BaseModel):
    """Главная строка оценки: эталонная ошибка и сопоставленная найденная ошибка."""

    project: str
    project_id: str
    criterion: str
    label: str
    gold_row_number: int
    gold_line_range: str
    gold_text: str
    found_finding_id: str | None = None
    found_checker: str | None = None
    found_line_range: str
    found_text: str
    match_type: str
    match_score: float
    counted: bool
    reason: str


class CriterionMetrics(BaseModel):
    """Метрики по одному критерию."""

    criterion: str
    label: str
    gold_total: int
    predicted_total: int
    true_positive: int
    false_positive: int
    false_negative: int
    precision: float
    recall: float
    f1_score: float


class PredictionSliceMetrics(BaseModel):
    """Метрики по срезу найденных ошибок: чекер, группа чекеров или действенный слой."""

    slice_name: str
    label: str
    predicted_total: int
    true_positive: int
    false_positive: int
    precision: float
    false_positive_share: float


class CostQualityMetrics(BaseModel):
    """Связка стоимости модельных проверок с качеством результата."""

    model_calls: int
    cache_hits: int
    total_tokens: int
    cost_usd: float
    cost_per_gold_true_positive: float | None = None
    cost_per_prediction: float | None = None
    cost_per_actionable_true_positive: float | None = None


class FalseNegativeAnalysisItem(BaseModel):
    """Одна пропущенная эталонная ошибка с причиной пропуска и следующим шагом."""

    project: str
    project_id: str
    criterion: str
    label: str
    gold_line_range: str
    gold_text: str
    nearest_finding_id: str | None = None
    nearest_checker: str | None = None
    nearest_match_type: str
    nearest_score: float
    reason_code: str
    reason_label: str
    next_step: str


class CorpusEvaluationSummary(BaseModel):
    """Итог оценки по корпусу проектов."""

    evaluated_criteria: list[str]
    gold_total: int
    predicted_total: int
    true_positive: int
    false_positive: int
    false_negative: int
    precision: float
    recall: float
    f1_score: float
    macro_precision: float
    macro_recall: float
    macro_f1_score: float
    overview_gold_total: int
    overview_predicted_total: int
    overview_true_positive: int
    overview_false_positive: int
    overview_false_negative: int
    overview_precision: float
    overview_recall: float
    overview_f1_score: float
    overview_macro_precision: float
    overview_macro_recall: float
    overview_macro_f1_score: float
    gold_scope_predicted_total: int
    gold_scope_true_positive: int
    gold_scope_false_positive: int
    gold_scope_false_negative: int
    gold_scope_precision: float
    gold_scope_recall: float
    gold_scope_f1_score: float
    gold_scope_macro_precision: float
    gold_scope_macro_recall: float
    gold_scope_macro_f1_score: float
    per_criterion: list[CriterionMetrics]
    overview_per_criterion: list[CriterionMetrics]
    checker_metrics: list[PredictionSliceMetrics] = Field(default_factory=list)
    checker_group_metrics: list[PredictionSliceMetrics] = Field(default_factory=list)
    actionable_metrics: PredictionSliceMetrics | None = None
    cost_quality: CostQualityMetrics | None = None
    false_negative_reason_counts: dict[str, int] = Field(default_factory=dict)
    false_negative_analysis: list[FalseNegativeAnalysisItem] = Field(default_factory=list)
    gold_items: list[GoldCorpusItem]
    gold_cases: list[GoldCorpusCase]
    matches: list[CorpusEvaluationMatch]
    detailed_false_positive_items: list[PredictedCorpusItem]
    false_positive_items: list[CorpusEvaluationKey]
    false_negative_items: list[CorpusEvaluationKey]
    project_mapping: dict[str, str]
    notes: list[str] = Field(default_factory=list)


@dataclass(frozen=True)
class _ProjectCandidate:
    project_id: str
    raw_name: str
    normalized_name: str
    tokens: frozenset[str]


@dataclass(frozen=True)
class _MatchCandidate:
    gold_case_id: str
    prediction_id: str
    match_type: str
    score: float
    reason: str


def evaluate_corpus_report(
    report: AuditReport,
    gold_xlsx_path: Path,
    *,
    matcher: str = "strict",
    judge_backend: str = "offline",
    judge_model: str | None = None,
    judge_api_key: str | None = None,
    judge_topk: int = 6,
    judge_cache_path: str | None = None,
    defects_only: bool = False,
    confidence_floor: float = 0.0,
    mirror_dedupe: bool = False,
    cap_repetitive: int = 0,
) -> CorpusEvaluationSummary:
    """Сравнивает отчёт аудита с Excel-разметкой на уровне конкретных эталонных ошибок."""

    unit_candidates = _project_candidates_from_report(report)
    units_by_id = {unit.unit_id: unit for unit in report.units}
    from content_factory.audit import gold_atomic

    if gold_atomic.is_atomic(gold_xlsx_path):
        gold_items, gold_cases, opinion_case_ids, mapping_notes = gold_atomic.load_atomic(
            gold_xlsx_path, unit_candidates
        )
    else:
        gold_items, mapping_notes = load_gold_items(gold_xlsx_path, unit_candidates)
        gold_cases = _gold_cases_from_items(gold_items)
        opinion_case_ids = set()
    if defects_only:
        if opinion_case_ids:
            gold_cases = [case for case in gold_cases if case.case_id not in opinion_case_ids]
        else:
            from content_factory.audit.aligner import is_opinion

            gold_cases = [case for case in gold_cases if not is_opinion(case.gold_text)]
    predicted_items = _predicted_items_from_report(report, units_by_id)
    evaluated_criteria = sorted({item.criterion for item in gold_cases})
    predicted_items_in_scope = [
        item
        for item in predicted_items
        if item.criterion in evaluated_criteria and _is_strict_evaluation_signal(item)
    ]
    if confidence_floor > 0.0 or mirror_dedupe or cap_repetitive > 0 or matcher == "anchor_judge":
        from content_factory.audit import aligner

        predicted_items_in_scope = aligner.confidence_gate(predicted_items_in_scope, confidence_floor)
        if mirror_dedupe:
            predicted_items_in_scope = aligner.dedupe_mirror(predicted_items_in_scope)
        if cap_repetitive > 0:
            predicted_items_in_scope, _capped = aligner.cap_repetitive(
                predicted_items_in_scope, per_issue_type=cap_repetitive
            )
    if matcher == "anchor_judge":
        from content_factory.audit import aligner

        judge = aligner.build_judge(
            judge_backend,
            api_key=judge_api_key,
            model=judge_model,
            cache_path=judge_cache_path,
        )
        matches, matched_prediction_ids = aligner.match_anchor_judge(
            gold_cases, predicted_items_in_scope, judge, topk=judge_topk
        )
    else:
        matches, matched_prediction_ids = _match_gold_cases(gold_cases, predicted_items_in_scope)
    detailed_false_positive_items = [
        item for item in predicted_items_in_scope if item.finding_id not in matched_prediction_ids
    ]
    detailed_true_positive = sum(1 for item in matches if item.counted)
    detailed_false_negative = len(gold_cases) - detailed_true_positive
    detailed_false_positive = len(detailed_false_positive_items)
    per_criterion = _per_criterion_detail_metrics(gold_cases, predicted_items_in_scope, matches)
    actionable_metrics = _prediction_slice_metrics(
        "actionable",
        "Действенные находки",
        [item for item in predicted_items_in_scope if _is_actionable_prediction(item)],
        matched_prediction_ids,
        detailed_false_positive,
    )
    checker_metrics = _checker_metrics(predicted_items_in_scope, matched_prediction_ids, detailed_false_positive)
    checker_group_metrics = _checker_group_metrics(predicted_items_in_scope, matched_prediction_ids, detailed_false_positive)
    false_negative_analysis = _false_negative_analysis(matches)
    false_negative_reason_counts = _false_negative_reason_counts(false_negative_analysis)
    cost_quality = _cost_quality_metrics(report, detailed_true_positive, actionable_metrics)

    overview_gold_keys = {
        CorpusEvaluationKey(project_id=item.project_id, criterion=criterion)
        for item in gold_items
        for criterion in item.criteria
    }
    overview_predicted_keys = _predicted_keys_from_report(report)

    overview_true_positive_keys = overview_gold_keys & overview_predicted_keys
    overview_false_positive_keys = overview_predicted_keys - overview_gold_keys
    overview_false_negative_keys = overview_gold_keys - overview_predicted_keys
    overview_per_criterion = _per_criterion_metrics(overview_gold_keys, overview_predicted_keys)

    return CorpusEvaluationSummary(
        evaluated_criteria=evaluated_criteria,
        gold_total=len(gold_cases),
        predicted_total=len(predicted_items_in_scope),
        true_positive=detailed_true_positive,
        false_positive=detailed_false_positive,
        false_negative=detailed_false_negative,
        precision=_safe_ratio(detailed_true_positive, detailed_true_positive + detailed_false_positive),
        recall=_safe_ratio(detailed_true_positive, detailed_true_positive + detailed_false_negative),
        f1_score=_f1(detailed_true_positive, detailed_false_positive, detailed_false_negative),
        macro_precision=_mean([item.precision for item in per_criterion]),
        macro_recall=_mean([item.recall for item in per_criterion]),
        macro_f1_score=_mean([item.f1_score for item in per_criterion]),
        overview_gold_total=len(overview_gold_keys),
        overview_predicted_total=len(overview_predicted_keys),
        overview_true_positive=len(overview_true_positive_keys),
        overview_false_positive=len(overview_false_positive_keys),
        overview_false_negative=len(overview_false_negative_keys),
        overview_precision=_safe_ratio(
            len(overview_true_positive_keys),
            len(overview_true_positive_keys) + len(overview_false_positive_keys),
        ),
        overview_recall=_safe_ratio(
            len(overview_true_positive_keys),
            len(overview_true_positive_keys) + len(overview_false_negative_keys),
        ),
        overview_f1_score=_f1(
            len(overview_true_positive_keys),
            len(overview_false_positive_keys),
            len(overview_false_negative_keys),
        ),
        overview_macro_precision=_mean([item.precision for item in overview_per_criterion]),
        overview_macro_recall=_mean([item.recall for item in overview_per_criterion]),
        overview_macro_f1_score=_mean([item.f1_score for item in overview_per_criterion]),
        gold_scope_predicted_total=len(predicted_items_in_scope),
        gold_scope_true_positive=detailed_true_positive,
        gold_scope_false_positive=detailed_false_positive,
        gold_scope_false_negative=detailed_false_negative,
        gold_scope_precision=_safe_ratio(
            detailed_true_positive,
            detailed_true_positive + detailed_false_positive,
        ),
        gold_scope_recall=_safe_ratio(
            detailed_true_positive,
            detailed_true_positive + detailed_false_negative,
        ),
        gold_scope_f1_score=_f1(
            detailed_true_positive,
            detailed_false_positive,
            detailed_false_negative,
        ),
        gold_scope_macro_precision=_mean([item.precision for item in per_criterion]),
        gold_scope_macro_recall=_mean([item.recall for item in per_criterion]),
        gold_scope_macro_f1_score=_mean([item.f1_score for item in per_criterion]),
        per_criterion=per_criterion,
        overview_per_criterion=overview_per_criterion,
        checker_metrics=checker_metrics,
        checker_group_metrics=checker_group_metrics,
        actionable_metrics=actionable_metrics,
        cost_quality=cost_quality,
        false_negative_reason_counts=false_negative_reason_counts,
        false_negative_analysis=false_negative_analysis,
        gold_items=gold_items,
        gold_cases=gold_cases,
        matches=matches,
        detailed_false_positive_items=sorted(
            detailed_false_positive_items,
            key=lambda item: (item.project_id, item.criterion, item.line_start or 0, item.finding_id),
        ),
        false_positive_items=sorted(overview_false_positive_keys, key=lambda item: (item.project_id, item.criterion)),
        false_negative_items=sorted(overview_false_negative_keys, key=lambda item: (item.project_id, item.criterion)),
        project_mapping={item.raw_project: item.matched_project for item in gold_items},
        notes=[
            "Основное сравнение выполняется на уровне атомарных ошибок: проект, критерий, строка/диапазон и текст.",
            "В строгую метрику не входят диагностические строки: низкоуверенные unknown, непроверенные ссылки и общие ресурсные предупреждения без имени файла.",
            "Старая метрика проект × критерий сохранена только как обзорная в overview_* полях.",
            "Excel-разметка нормализуется эвристически из колонок 'Проблема' и 'Детали'.",
            *mapping_notes,
        ],
    )


def load_gold_items(
    gold_xlsx_path: Path,
    unit_candidates: list[_ProjectCandidate],
) -> tuple[list[GoldCorpusItem], list[str]]:
    """Читает Excel и переводит строки в эталонные критерии."""

    workbook = load_workbook(gold_xlsx_path, data_only=True)
    sheet = workbook.active
    header = _header_map(sheet)
    required_columns = {PROJECT_COLUMN, PROBLEM_COLUMN, DETAILS_COLUMN}
    missing_columns = sorted(required_columns - set(header))
    if missing_columns:
        raise ValueError(f"В Excel не найдены обязательные колонки: {', '.join(missing_columns)}")

    items: list[GoldCorpusItem] = []
    notes: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_project = _cell_text(row, header[PROJECT_COLUMN])
        raw_problem = _cell_text(row, header[PROBLEM_COLUMN])
        details = _cell_text(row, header[DETAILS_COLUMN])
        if not raw_project and not raw_problem and not details:
            continue

        candidate, score = _match_project(raw_project, unit_candidates)
        criteria = _criteria_from_gold_row(raw_problem, details)
        if not criteria:
            notes.append(f"Строка {row_number}: не удалось вывести критерий из типа проблемы {raw_problem!r}.")
            continue
        if score < 0.55:
            notes.append(
                f"Строка {row_number}: слабое сопоставление проекта {raw_project!r} "
                f"с папкой {candidate.raw_name!r}, score={score:.2f}."
            )

        items.append(
            GoldCorpusItem(
                row_number=row_number,
                raw_project=raw_project,
                matched_project=candidate.raw_name,
                project_id=candidate.project_id,
                raw_problem=raw_problem,
                details=details,
                criteria=criteria,
            )
        )
    return items, notes


def write_corpus_evaluation(summary: CorpusEvaluationSummary, output_dir: Path) -> None:
    """Записывает машинный и табличный отчёт оценки."""

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "corpus_evaluation.json").write_text(
        json.dumps(summary.model_dump(mode="json"), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    _write_matches_csv(summary.matches, output_dir / "corpus_evaluation_main.csv")
    _write_metrics_csv(summary, output_dir / "corpus_evaluation_by_criterion.csv")
    _write_overview_metrics_csv(summary, output_dir / "corpus_evaluation_overview_by_criterion.csv")
    _write_prediction_slice_metrics_csv(summary.checker_metrics, output_dir / "corpus_evaluation_by_checker.csv")
    _write_prediction_slice_metrics_csv(summary.checker_group_metrics, output_dir / "corpus_evaluation_by_checker_group.csv")
    _write_false_negative_analysis_csv(summary.false_negative_analysis, output_dir / "corpus_false_negative_analysis.csv")
    _write_detailed_false_positive_csv(summary.detailed_false_positive_items, output_dir / "corpus_false_positive_detailed.csv")
    _write_items_csv(summary.false_negative_items, output_dir / "corpus_false_negative.csv")
    _write_items_csv(summary.false_positive_items, output_dir / "corpus_false_positive.csv")


def _gold_cases_from_items(items: list[GoldCorpusItem]) -> list[GoldCorpusCase]:
    """Разбивает Excel-строки на атомарные эталонные случаи."""

    cases: list[GoldCorpusCase] = []
    for item in items:
        detail_cases = _split_gold_detail_cases(item.details or item.raw_problem)
        for index, detail in enumerate(detail_cases, start=1):
            detail_criteria = _criteria_from_gold_case(str(detail["text"])) or item.criteria
            for criterion in detail_criteria:
                cases.append(
                    GoldCorpusCase(
                        case_id=f"gold_{item.row_number}_{criterion}_{index}",
                        row_number=item.row_number,
                        raw_project=item.raw_project,
                        matched_project=item.matched_project,
                        project_id=item.project_id,
                        criterion=criterion,
                        line_start=detail["line_start"],
                        line_end=detail["line_end"],
                        gold_text=detail["text"],
                    )
                )
    return cases


def _criteria_from_gold_case(text: str) -> list[str]:
    """Выводит критерий из конкретной строки эталонной ошибки, без размножения на весь тип проблемы."""

    lowered = str(text or "").lower()
    markers: tuple[tuple[Criterion, tuple[str, ...]], ...] = (
        (
            Criterion.LINKS,
            (
                "сломанная ссылка",
                "ссылка",
                "url",
                "http",
                "https",
            ),
        ),
        (
            Criterion.TECHNOLOGY_FRESHNESS,
            (
                "устар",
                "неактуаль",
                "версии ос",
                "библиотек",
                "фреймворк",
                "стек",
            ),
        ),
        (
            Criterion.CHECKLIST_ALIGNMENT,
            (
                "чеклист",
                "чек-лист",
                "чек лист",
                "check-list",
                "checklist",
            ),
        ),
        (
            Criterion.READABILITY,
            (
                "опечат",
                "двоеточ",
                "нумерац",
                "пронумер",
                "граммат",
                "тавтолог",
                "формулиров",
                "кавыч",
                "input ",
                "output",
                "example",
                "result",
            ),
        ),
        (
            Criterion.CORRECTNESS,
            (
                "противореч",
                "некоррект",
                "неверн",
                "не является",
                "по факту",
                "ошибка в задании",
            ),
        ),
    )
    result: list[Criterion] = []
    for criterion, aliases in markers:
        if any(alias in lowered for alias in aliases):
            result.append(criterion)
    return [criterion.value for criterion in dict.fromkeys(result)]


def _split_gold_detail_cases(text: str) -> list[dict[str, Any]]:
    """Выделяет отдельные ошибки из многострочного описания разметки."""

    lines = _normalise_gold_detail_lines(text)
    if not lines:
        lines = [str(text or "").strip()]
    cases: list[dict[str, Any]] = []
    for line in lines:
        line_start, line_end = _line_range_from_text(line)
        cases.append(
            {
                "line_start": line_start,
                "line_end": line_end,
                "text": _strip_gold_list_marker(line),
            }
        )
    if cases:
        return cases
    line_start, line_end = _line_range_from_text(text)
    return [{"line_start": line_start, "line_end": line_end, "text": str(text or "").strip()}]


def _normalise_gold_detail_lines(text: str) -> list[str]:
    """Оставляет только строки с дефектами, отделяя решения, доказательства и шапки таблиц."""

    result: list[str] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip().strip('"').strip()
        if not line or _looks_like_gold_metadata_line(line):
            continue
        if _starts_non_defect_gold_section(line):
            break
        result.append(line)
    return result


def _looks_like_gold_metadata_line(value: str) -> bool:
    """Отсекает служебные строки из Excel-описания, если они попали в детали."""

    lowered = re.sub(r"\s+", " ", value.lower().strip(" :"))
    exact_headings = {
        "строка - проблема - решение",
        "строка проблема решение",
        "проблема - решение",
        "problem - solution",
        "line - problem - solution",
    }
    if lowered in exact_headings:
        return True
    return lowered.startswith(("тип проблемы", "проект", "критерий", "архив", ":small_blue_diamond: description"))


def _starts_non_defect_gold_section(value: str) -> bool:
    """Определяет начало раздела с решением или доказательствами, а не с эталонной ошибкой."""

    lowered = re.sub(r"\s+", " ", value.lower().strip(" :"))
    prefixes = (
        "предложение по решению",
        "предлагаемое решение",
        "решение",
        "аргументы и исследования",
        "аргументы",
        "исследования",
        "доказательства",
        "подтверждение",
        "скрины",
        "скриншоты",
    )
    return lowered.startswith(prefixes)


def _strip_gold_list_marker(value: str) -> str:
    """Убирает маркер списка, не путая его с номером строки исходного файла."""

    text = str(value or "").strip()
    bullet_match = re.match(r"^\s*[*•]\s*(.+)$", text)
    if bullet_match is not None:
        return bullet_match.group(1).strip()
    ordered_match = re.match(r"^\s*\d{1,3}[.)]\s+(.+)$", text)
    if ordered_match is not None:
        return ordered_match.group(1).strip()
    return text


def _line_range_from_text(value: str) -> tuple[int | None, int | None]:
    """Достаёт строку или диапазон строк из эталонного описания."""

    text = str(value or "")
    patterns = (
        r"^\s*(?:строк[аеи]\s*)?(\d{1,5})\s*[–—-]\s*(\d{1,5})\s+[–—-]",
        r"^\s*(?:строк[аеи]\s*)?(\d{1,5})\s+[–—-]",
        r"\bстрок[аеи]?\s*(\d{1,5})\s*[–—-]\s*(\d{1,5})\b",
        r"\bстрок[аеи]?\s*(\d{1,5})\b(?=\s*[:.)]|\s|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match is None:
            continue
        start = int(match.group(1))
        end = int(match.group(2)) if match.lastindex and match.lastindex >= 2 and match.group(2) else start
        return min(start, end), max(start, end)
    return None, None


def _predicted_items_from_report(report: AuditReport, units_by_id: dict[str, Any]) -> list[PredictedCorpusItem]:
    """Преобразует Findings в атомарные найденные случаи для детальной оценки."""

    items: list[PredictedCorpusItem] = []
    for finding in report.findings:
        unit = units_by_id.get(finding.unit_id)
        if unit is None:
            continue
        location = finding.location
        items.append(
            PredictedCorpusItem(
                finding_id=finding.finding_id,
                project_id=finding.unit_id,
                project=unit.name,
                criterion=finding.criterion.value,
                checker_name=finding.checker_name,
                line_start=location.line_start if location else None,
                line_end=location.line_end if location else None,
                file_path=location.file_path if location else None,
                severity=finding.severity.value,
                verdict=finding.verdict.value,
                confidence=finding.confidence,
                issue_type=str(finding.extra.get("issue_type") or ""),
                found_text=_finding_text(finding),
            )
        )
    return items


def _is_strict_evaluation_signal(item: PredictedCorpusItem) -> bool:
    """Оставляет в строгой метрике только находки, похожие на проверяемый дефект."""

    if item.verdict == "pass":
        return False
    if item.verdict == "unknown" and (item.confidence or 0.0) < 0.8:
        return False
    if item.checker_name == "link_checker":
        return item.verdict == "fail"
    if item.checker_name == "resource_availability_checker":
        return item.issue_type in {"missing_local_resource", "unconfirmed_environment_path"}
    return True


def _finding_text(finding: Finding) -> str:
    """Собирает человекочитаемый текст найденной ошибки из цитаты, основания и рекомендации."""

    parts: list[str] = []
    if finding.quote:
        parts.append(str(finding.quote))
    for evidence in finding.evidence[:2]:
        if evidence.detail:
            parts.append(evidence.detail)
    if finding.recommendation:
        parts.append(finding.recommendation)
    issue_type = finding.extra.get("issue_type")
    if issue_type:
        parts.append(str(issue_type))
    return " | ".join(part.strip() for part in parts if part and part.strip())


def _match_gold_cases(
    gold_cases: list[GoldCorpusCase],
    predicted_items: list[PredictedCorpusItem],
) -> tuple[list[CorpusEvaluationMatch], set[str]]:
    """Сопоставляет эталонные и найденные ошибки один-к-одному."""

    gold_by_id = {item.case_id: item for item in gold_cases}
    predicted_by_id = {item.finding_id: item for item in predicted_items}
    counted_candidates: list[_MatchCandidate] = []
    best_any_by_gold: dict[str, _MatchCandidate] = {}
    for gold in gold_cases:
        for predicted in predicted_items:
            if gold.project_id != predicted.project_id or gold.criterion != predicted.criterion:
                continue
            candidate = _score_gold_prediction_match(gold, predicted)
            current_best = best_any_by_gold.get(gold.case_id)
            if current_best is None or candidate.score > current_best.score:
                best_any_by_gold[gold.case_id] = candidate
            if _is_counted_match(candidate):
                counted_candidates.append(candidate)

    assigned_gold: set[str] = set()
    assigned_predictions: set[str] = set()
    assigned_by_gold: dict[str, _MatchCandidate] = {}
    for candidate in sorted(counted_candidates, key=lambda item: item.score, reverse=True):
        if candidate.gold_case_id in assigned_gold or candidate.prediction_id in assigned_predictions:
            continue
        assigned_gold.add(candidate.gold_case_id)
        assigned_predictions.add(candidate.prediction_id)
        assigned_by_gold[candidate.gold_case_id] = candidate

    rows: list[CorpusEvaluationMatch] = []
    for gold in gold_cases:
        candidate = assigned_by_gold.get(gold.case_id)
        counted = candidate is not None
        if candidate is None:
            candidate = _best_unassigned_candidate(gold, predicted_items, assigned_predictions)
        prediction = predicted_by_id.get(candidate.prediction_id) if candidate is not None else None
        rows.append(_match_row(gold, prediction, candidate, counted))
    return rows, assigned_predictions


def _best_unassigned_candidate(
    gold: GoldCorpusCase,
    predicted_items: list[PredictedCorpusItem],
    assigned_predictions: set[str],
) -> _MatchCandidate | None:
    """Выбирает лучший неиспользованный прогноз для объяснения незачёта."""

    best: _MatchCandidate | None = None
    for predicted in predicted_items:
        if predicted.finding_id in assigned_predictions:
            continue
        if gold.project_id != predicted.project_id or gold.criterion != predicted.criterion:
            continue
        candidate = _score_gold_prediction_match(gold, predicted)
        if best is None or candidate.score > best.score:
            best = candidate
    return best


def _score_gold_prediction_match(gold: GoldCorpusCase, predicted: PredictedCorpusItem) -> _MatchCandidate:
    """Оценивает пару эталонная ошибка ↔ найденная ошибка."""

    line_relation = _line_relation(gold.line_start, gold.line_end, predicted.line_start, predicted.line_end)
    text_score = _text_match_score(gold.gold_text, predicted.found_text)
    if _same_missing_artifact_signal(gold.gold_text, predicted.found_text):
        return _candidate(
            gold,
            predicted,
            "artifact_missing_signal",
            max(0.78, text_score),
            "Совпали проект, критерий, ожидаемый маркер, тип артефакта и факт отсутствия маркера.",
        )
    if line_relation == "overlap" and text_score >= 0.25:
        return _candidate(gold, predicted, "line_and_text", max(0.9, text_score), "Совпали проект, критерий, диапазон строк и ключевой текст ошибки.")
    if line_relation == "overlap":
        return _candidate(gold, predicted, "line_overlap", 0.82, "Совпали проект, критерий и строка/диапазон; текстовое совпадение слабее, но строка указывает на тот же дефект.")
    if line_relation == "near" and text_score >= 0.2:
        return _candidate(gold, predicted, "near_line_and_text", max(0.75, text_score), "Строки отличаются не более чем на две, критерий совпал, текст ошибки похож.")
    if text_score >= 0.55:
        return _candidate(gold, predicted, "text_similarity", text_score, "Совпали проект, критерий и текст ошибки; строка отсутствует или отличается.")
    return _candidate(gold, predicted, "criterion_only", max(0.15, text_score), "Совпал только проект и критерий; для основной метрики это не засчитывается.")


def _candidate(
    gold: GoldCorpusCase,
    predicted: PredictedCorpusItem,
    match_type: str,
    score: float,
    reason: str,
) -> _MatchCandidate:
    """Создаёт внутренний объект кандидата сопоставления."""

    return _MatchCandidate(
        gold_case_id=gold.case_id,
        prediction_id=predicted.finding_id,
        match_type=match_type,
        score=round(min(max(score, 0.0), 1.0), 4),
        reason=reason,
    )


def _is_counted_match(candidate: _MatchCandidate) -> bool:
    """Решает, засчитывается ли совпадение в основной метрике."""

    return candidate.match_type in {
        "line_and_text",
        "line_overlap",
        "near_line_and_text",
        "text_similarity",
        "artifact_missing_signal",
    }


def _match_row(
    gold: GoldCorpusCase,
    prediction: PredictedCorpusItem | None,
    candidate: _MatchCandidate | None,
    counted: bool,
) -> CorpusEvaluationMatch:
    """Создаёт строку главного отчёта сопоставления."""

    if prediction is None or candidate is None:
        return CorpusEvaluationMatch(
            project=gold.matched_project,
            project_id=gold.project_id,
            criterion=gold.criterion,
            label=_criterion_label(gold.criterion),
            gold_row_number=gold.row_number,
            gold_line_range=_format_range(gold.line_start, gold.line_end),
            gold_text=gold.gold_text,
            found_line_range="",
            found_text="",
            match_type="missed",
            match_score=0.0,
            counted=False,
            reason="По этому проекту и критерию не найдено подходящей ошибки с совпадающей строкой или текстом.",
        )
    return CorpusEvaluationMatch(
        project=gold.matched_project,
        project_id=gold.project_id,
        criterion=gold.criterion,
        label=_criterion_label(gold.criterion),
        gold_row_number=gold.row_number,
        gold_line_range=_format_range(gold.line_start, gold.line_end),
        gold_text=gold.gold_text,
        found_finding_id=prediction.finding_id,
        found_checker=prediction.checker_name,
        found_line_range=_format_prediction_range(prediction),
        found_text=prediction.found_text,
        match_type=candidate.match_type,
        match_score=candidate.score,
        counted=counted,
        reason=candidate.reason if counted else f"{candidate.reason} Найденная ошибка показана для разбора, но не засчитана.",
    )


def _line_relation(
    gold_start: int | None,
    gold_end: int | None,
    pred_start: int | None,
    pred_end: int | None,
) -> str:
    """Определяет отношение диапазонов строк."""

    if gold_start is None or pred_start is None:
        return "none"
    gold_end = gold_end or gold_start
    pred_end = pred_end or pred_start
    if gold_start <= pred_end and pred_start <= gold_end:
        return "overlap"
    distance = min(abs(gold_start - pred_end), abs(pred_start - gold_end))
    return "near" if distance <= 2 else "far"


def _text_match_score(gold_text: str, found_text: str) -> float:
    """Считает похожесть текста эталона и найденной ошибки."""

    gold = _normalize_match_text(gold_text)
    found = _normalize_match_text(found_text)
    if not gold or not found:
        return 0.0
    if gold in found or found in gold:
        return 0.95
    gold_tokens = set(gold.split())
    found_tokens = set(found.split())
    overlap = len(gold_tokens & found_tokens)
    token_score = overlap / max(min(len(gold_tokens), len(found_tokens)), 1)
    sequence_score = SequenceMatcher(a=gold, b=found).ratio()
    return round(max(token_score, sequence_score), 4)


def _same_missing_artifact_signal(gold_text: str, found_text: str) -> bool:
    """Сопоставляет формулировки об отсутствующем маркере внутри артефакта."""

    gold = _normalize_match_text(gold_text)
    found = _normalize_match_text(found_text)
    if not gold or not found:
        return False
    shared_commands = _artifact_command_markers(gold) & _artifact_command_markers(found)
    if not shared_commands:
        return False
    return _mentions_artifact(gold) and _mentions_artifact(found) and _mentions_absence(gold) and _mentions_absence(found)


def _artifact_command_markers(text: str) -> set[str]:
    """Находит командные маркеры, важные для сравнения ошибок по артефактам."""

    commands = {"whoami", "id", "uname", "ls", "hostname", "ifconfig", "ipconfig", "tcpdump", "tshark"}
    return {command for command in commands if re.search(rf"(?<![\w-]){re.escape(command)}(?![\w-])", text)}


def _mentions_artifact(text: str) -> bool:
    """Проверяет, что формулировка говорит об артефакте, дампе или захвате."""

    return bool(re.search(r"\b(pcap|pcapng|dump|capture|trace|log)\b|дамп|захват|артефакт", text))


def _mentions_absence(text: str) -> bool:
    """Проверяет отрицание или отсутствие в формулировке дефекта."""

    return bool(re.search(r"\bнет\b|не найден|не содержит|отсутств|missing|not found|without", text))


def _normalize_match_text(value: str) -> str:
    """Нормализует текст для сравнения эталона с находкой."""

    text = str(value or "").lower()
    text = re.sub(r"https?:[/\\]+", " ", text)
    text = re.sub(r"[`*_\"'«»()\[\]{}:;,.!?/\\|]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _format_range(start: int | None, end: int | None) -> str:
    """Форматирует диапазон строк для отчёта."""

    if start is None:
        return ""
    if end is None or end == start:
        return str(start)
    return f"{start}-{end}"


def _format_prediction_range(item: PredictedCorpusItem) -> str:
    """Форматирует файл и строку найденной ошибки."""

    line_range = _format_range(item.line_start, item.line_end)
    if item.file_path and line_range:
        return f"{item.file_path}:{line_range}"
    return item.file_path or line_range


def _per_criterion_detail_metrics(
    gold_cases: list[GoldCorpusCase],
    predicted_items: list[PredictedCorpusItem],
    matches: list[CorpusEvaluationMatch],
) -> list[CriterionMetrics]:
    """Считает основную метрику по критериям на уровне атомарных ошибок."""

    criteria = sorted({item.criterion for item in gold_cases} | {item.criterion for item in predicted_items})
    matched_prediction_ids = {item.found_finding_id for item in matches if item.counted and item.found_finding_id}
    metrics: list[CriterionMetrics] = []
    for criterion in criteria:
        gold = [item for item in gold_cases if item.criterion == criterion]
        predicted = [item for item in predicted_items if item.criterion == criterion]
        tp = sum(1 for item in matches if item.criterion == criterion and item.counted)
        matched_predicted = [item for item in predicted if item.finding_id in matched_prediction_ids]
        fp = len(predicted) - len(matched_predicted)
        fn = len(gold) - tp
        metrics.append(
            CriterionMetrics(
                criterion=criterion,
                label=_criterion_label(criterion),
                gold_total=len(gold),
                predicted_total=len(predicted),
                true_positive=tp,
                false_positive=fp,
                false_negative=fn,
                precision=_safe_ratio(tp, tp + fp),
                recall=_safe_ratio(tp, tp + fn),
                f1_score=_f1(tp, fp, fn),
            )
        )
    return metrics


def _checker_metrics(
    predicted_items: list[PredictedCorpusItem],
    matched_prediction_ids: set[str],
    total_false_positive: int,
) -> list[PredictionSliceMetrics]:
    """Считает precision по каждому чекеру, у которого были находки в gold-scope."""

    checker_names = sorted({item.checker_name for item in predicted_items})
    return [
        _prediction_slice_metrics(
            checker,
            checker,
            [item for item in predicted_items if item.checker_name == checker],
            matched_prediction_ids,
            total_false_positive,
        )
        for checker in checker_names
    ]


def _checker_group_metrics(
    predicted_items: list[PredictedCorpusItem],
    matched_prediction_ids: set[str],
    total_false_positive: int,
) -> list[PredictionSliceMetrics]:
    """Считает precision по группам чекеров, чтобы не смешивать разные типы качества."""

    groups = sorted({_checker_group(item.checker_name) for item in predicted_items})
    return [
        _prediction_slice_metrics(
            group,
            CHECKER_GROUP_LABELS.get(group, group),
            [item for item in predicted_items if _checker_group(item.checker_name) == group],
            matched_prediction_ids,
            total_false_positive,
        )
        for group in groups
    ]


def _prediction_slice_metrics(
    slice_name: str,
    label: str,
    predicted_items: list[PredictedCorpusItem],
    matched_prediction_ids: set[str],
    total_false_positive: int,
) -> PredictionSliceMetrics:
    """Считает точность для произвольного среза предсказаний."""

    true_positive = sum(1 for item in predicted_items if item.finding_id in matched_prediction_ids)
    false_positive = len(predicted_items) - true_positive
    return PredictionSliceMetrics(
        slice_name=slice_name,
        label=label,
        predicted_total=len(predicted_items),
        true_positive=true_positive,
        false_positive=false_positive,
        precision=_safe_ratio(true_positive, true_positive + false_positive),
        false_positive_share=_safe_ratio(false_positive, total_false_positive),
    )


def _checker_group(checker_name: str) -> str:
    """Возвращает устойчивую группу чекера для продуктовых метрик."""

    return CHECKER_GROUPS.get(checker_name, "other")


def _is_actionable_prediction(item: PredictedCorpusItem) -> bool:
    """Выделяет находки, которые должны попадать в рабочий фокус методолога."""

    if item.verdict in {"pass", "unknown"}:
        return False
    if item.severity in {"critical", "major"}:
        return True
    if item.severity == "minor":
        return _checker_group(item.checker_name) not in {"methodology", "factcheck"}
    return False


def _cost_quality_metrics(
    report: AuditReport,
    true_positive: int,
    actionable_metrics: PredictionSliceMetrics,
) -> CostQualityMetrics:
    """Связывает стоимость модельных вызовов с количеством подтверждённых находок."""

    usage = report.summary.model_usage
    return CostQualityMetrics(
        model_calls=usage.calls_total,
        cache_hits=usage.cache_hits,
        total_tokens=usage.total_tokens,
        cost_usd=round(float(usage.cost_usd), 6),
        cost_per_gold_true_positive=_safe_money_ratio(float(usage.cost_usd), true_positive),
        cost_per_prediction=_safe_money_ratio(float(usage.cost_usd), report.summary.findings_total),
        cost_per_actionable_true_positive=_safe_money_ratio(float(usage.cost_usd), actionable_metrics.true_positive),
    )


def _false_negative_analysis(matches: list[CorpusEvaluationMatch]) -> list[FalseNegativeAnalysisItem]:
    """Классифицирует пропуски по причине, чтобы превращать FN в план работ."""

    result: list[FalseNegativeAnalysisItem] = []
    for match in matches:
        if match.counted:
            continue
        reason_code, reason_label, next_step = _classify_false_negative(match)
        result.append(
            FalseNegativeAnalysisItem(
                project=match.project,
                project_id=match.project_id,
                criterion=match.criterion,
                label=match.label,
                gold_line_range=match.gold_line_range,
                gold_text=match.gold_text,
                nearest_finding_id=match.found_finding_id,
                nearest_checker=match.found_checker,
                nearest_match_type=match.match_type,
                nearest_score=match.match_score,
                reason_code=reason_code,
                reason_label=reason_label,
                next_step=next_step,
            )
        )
    return result


def _false_negative_reason_counts(items: list[FalseNegativeAnalysisItem]) -> dict[str, int]:
    """Считает распределение причин пропусков."""

    counts: dict[str, int] = defaultdict(int)
    for item in items:
        counts[item.reason_code] += 1
    return dict(sorted(counts.items()))


def _classify_false_negative(match: CorpusEvaluationMatch) -> tuple[str, str, str]:
    """Даёт инженерную причину пропуска для одного эталонного случая."""

    text = _normalize_match_text(match.gold_text)
    if _is_meaningful_near_miss(match):
        return (
            "near_miss_matching",
            "Похожая находка была, но строгий матч её не засчитал",
            "Проверить пороги сопоставления или критерий/строку у найденной ошибки.",
        )
    if _looks_like_deterministic_missing(text, match.criterion):
        return (
            "deterministic_possible",
            "Можно ловить правилом",
            "Добавить или расширить детерминированный чекер с тестом на этот паттерн.",
        )
    if _needs_external_data(text, match.criterion):
        return (
            "needs_external_data",
            "Нужен внешний источник данных",
            "Подключить метаданные платформы, манифест, статистику прохождения или курируемую базу.",
        )
    if _looks_like_model_semantics(text, match.criterion):
        return (
            "model_possible",
            "Нужен смысловой анализ",
            "Добавить модельный/семантический слой с проверяемым JSON-контрактом и примерами.",
        )
    return (
        "out_of_scope_or_needs_review",
        "Требуется ручной разбор области применимости",
        "Решить, должна ли локальная проверка ловить этот класс ошибок.",
    )


def _is_meaningful_near_miss(match: CorpusEvaluationMatch) -> bool:
    """Отличает близкий промах матчинга от случайной находки того же критерия."""

    if not match.found_finding_id or match.match_type == "missed":
        return False
    if match.match_type in {"line_overlap", "near_line_and_text", "text_similarity", "artifact_missing_signal"}:
        return True
    return match.match_type == "criterion_only" and match.match_score >= 0.45


def _looks_like_deterministic_missing(text: str, criterion: str) -> bool:
    """Определяет пропуски, которые стоит закрывать правилами."""

    deterministic_markers = (
        "https",
        "http",
        "ссылка",
        "url",
        "двоеточ",
        "нумерац",
        "пронумер",
        "кавыч",
        "опечат",
        "тафтолог",
        "тавтолог",
        "input",
        "output",
        "example",
        "result",
    )
    if any(marker in text for marker in deterministic_markers):
        return True
    return criterion == Criterion.READABILITY.value and bool(re.search(r"\b\d{1,5}\b", text))


def _needs_external_data(text: str, criterion: str) -> bool:
    """Определяет пропуски, которые нельзя честно закрыть локальным анализом."""

    markers = (
        "платформ",
        "репозитор",
        "обновлен",
        "обновл",
        "6мес",
        "год",
        "экзамен",
        "финаль",
        "время прохождения",
        "попыт",
        "доступн на территории",
        "рф",
        "лиценз",
        "права",
    )
    return criterion in {Criterion.ACTUALITY.value, Criterion.RIGHTS.value, Criterion.EXAM.value} and any(
        marker in text for marker in markers
    )


def _looks_like_model_semantics(text: str, criterion: str) -> bool:
    """Определяет пропуски, где нужна смысловая проверка, а не простое правило."""

    semantic_markers = (
        "противореч",
        "не является",
        "некоррект",
        "неверн",
        "по факту",
        "чеклист",
        "чек лист",
        "checklist",
        "check list",
        "артефакт",
        "ожида",
        "требован",
        "бизнес",
        "рын",
    )
    return criterion in {
        Criterion.CORRECTNESS.value,
        Criterion.CHECKLIST_ALIGNMENT.value,
        Criterion.MARKET_FIT.value,
    } or any(marker in text for marker in semantic_markers)


def _project_candidates_from_report(report: AuditReport) -> list[_ProjectCandidate]:
    """Создаёт кандидаты сопоставления из единиц отчёта."""

    return [
        _ProjectCandidate(
            project_id=unit.unit_id,
            raw_name=unit.name,
            normalized_name=_normalize_project_name(unit.name),
            tokens=frozenset(_project_tokens(unit.name)),
        )
        for unit in report.units
    ]


def _predicted_keys_from_report(report: AuditReport) -> set[CorpusEvaluationKey]:
    """Берёт все критерии, по которым алгоритм нашёл хотя бы один случай."""

    unit_ids = {unit.unit_id for unit in report.units}
    result: set[CorpusEvaluationKey] = set()
    for finding in report.findings:
        if finding.unit_id not in unit_ids:
            continue
        result.add(CorpusEvaluationKey(project_id=finding.unit_id, criterion=finding.criterion.value))
    return result


def _header_map(sheet: Any) -> dict[str, int]:
    """Возвращает индексы колонок по первой строке Excel."""

    result: dict[str, int] = {}
    for index, cell in enumerate(sheet[1]):
        if cell.value:
            result[str(cell.value).strip()] = index
    return result


def _cell_text(row: tuple[Any, ...], index: int) -> str:
    """Безопасно достаёт текст ячейки."""

    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _match_project(raw_project: str, candidates: list[_ProjectCandidate]) -> tuple[_ProjectCandidate, float]:
    """Сопоставляет короткое имя проекта из Excel с фактической папкой."""

    normalized = _normalize_project_name(raw_project)
    tokens = set(_project_tokens(raw_project))
    best: tuple[_ProjectCandidate, float] | None = None
    for candidate in candidates:
        score = _project_match_score(normalized, tokens, candidate)
        if best is None or score > best[1]:
            best = (candidate, score)
    if best is None:
        raise ValueError("В отчёте нет единиц контента для сопоставления с Excel.")
    return best


def _project_match_score(normalized: str, tokens: set[str], candidate: _ProjectCandidate) -> float:
    """Считает устойчивый score для разных написаний имени проекта."""

    if not normalized:
        return 0.0
    if normalized == candidate.normalized_name:
        return 1.0
    if normalized in candidate.normalized_name or candidate.normalized_name in normalized:
        return 0.95
    token_overlap = len(tokens & set(candidate.tokens)) / max(len(tokens), 1)
    sequence_score = SequenceMatcher(a=normalized, b=candidate.normalized_name).ratio()
    prefix_score = 0.0
    if tokens and candidate.tokens and next(iter(tokens)) in candidate.tokens:
        prefix_score = 0.25
    return max(sequence_score, token_overlap + prefix_score)


def _criteria_from_gold_row(raw_problem: str, details: str) -> list[str]:
    """Выводит наши критерии из свободного описания проблемы."""

    # The gold sheet already states the criterion in the "Проблема" column;
    # trust an exact label match before falling back to fuzzy alias matching.
    label_to_criterion = {label.lower(): criterion for criterion, label in CRITERION_LABELS.items()}
    direct = label_to_criterion.get(raw_problem.strip().lower())
    if direct is not None and direct != Criterion.ACTUALITY:
        return [direct.value]

    problem_text = raw_problem.lower()
    detail_text = details.lower()
    criteria: list[Criterion] = []

    # Тип проблемы надёжнее деталей, поэтому сначала используем его.
    for criterion, aliases in CRITERION_ALIASES.items():
        if any(alias in problem_text for alias in aliases):
            criteria.append(criterion)

    # Детали добавляют критерии только по сильным маркерам, чтобы не раздувать эталон.
    detail_markers: dict[Criterion, tuple[str, ...]] = {
        Criterion.LINKS: ("сломанная ссылка", "битая ссылка", "неверная ссылка", "url", "http", "https"),
        Criterion.TECHNOLOGY_FRESHNESS: ("неактуаль", "устар", "старый стандарт", "версии ос", "библиотек", "стек"),
        Criterion.FACTS: ("факт", "определение", "пример", "таблица вывода", "расхождение"),
        Criterion.CHECKLIST_ALIGNMENT: ("в чеклисте", "в чек-листе", "чеклист", "чек лист"),
        Criterion.READABILITY: ("опечат", "нумерация", "пронумер", "грамматика"),
        Criterion.CORRECTNESS: ("противореч", "некоррект", "по факту", "не является", "отсутствует"),
    }
    for criterion, markers in detail_markers.items():
        if any(marker in detail_text for marker in markers):
            criteria.append(criterion)

    return [criterion.value for criterion in dict.fromkeys(criteria)]


def _per_criterion_metrics(
    gold_keys: set[CorpusEvaluationKey],
    predicted_keys: set[CorpusEvaluationKey],
) -> list[CriterionMetrics]:
    """Считает метрики по каждому критерию, который есть в эталоне или прогнозе."""

    criteria = sorted({item.criterion for item in gold_keys | predicted_keys})
    metrics: list[CriterionMetrics] = []
    for criterion in criteria:
        gold = {item for item in gold_keys if item.criterion == criterion}
        predicted = {item for item in predicted_keys if item.criterion == criterion}
        tp = len(gold & predicted)
        fp = len(predicted - gold)
        fn = len(gold - predicted)
        metrics.append(
            CriterionMetrics(
                criterion=criterion,
                label=_criterion_label(criterion),
                gold_total=len(gold),
                predicted_total=len(predicted),
                true_positive=tp,
                false_positive=fp,
                false_negative=fn,
                precision=_safe_ratio(tp, tp + fp),
                recall=_safe_ratio(tp, tp + fn),
                f1_score=_f1(tp, fp, fn),
            )
        )
    return metrics


def _write_metrics_csv(summary: CorpusEvaluationSummary, output_path: Path) -> None:
    """Пишет основные детальные метрики по критериям в CSV."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "criterion",
                "label",
                "gold_total",
                "predicted_total",
                "true_positive",
                "false_positive",
                "false_negative",
                "precision",
                "recall",
                "f1_score",
            ],
        )
        writer.writeheader()
        for item in summary.per_criterion:
            writer.writerow(item.model_dump(mode="json"))


def _write_overview_metrics_csv(summary: CorpusEvaluationSummary, output_path: Path) -> None:
    """Пишет старую обзорную метрику проект × критерий в отдельный CSV."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "criterion",
                "label",
                "gold_total",
                "predicted_total",
                "true_positive",
                "false_positive",
                "false_negative",
                "precision",
                "recall",
                "f1_score",
            ],
        )
        writer.writeheader()
        for item in summary.overview_per_criterion:
            writer.writerow(item.model_dump(mode="json"))


def _write_prediction_slice_metrics_csv(items: list[PredictionSliceMetrics], output_path: Path) -> None:
    """Пишет метрики по чекерам или группам чекеров."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "slice_name",
                "label",
                "predicted_total",
                "true_positive",
                "false_positive",
                "precision",
                "false_positive_share",
            ],
        )
        writer.writeheader()
        for item in items:
            writer.writerow(item.model_dump(mode="json"))


def _write_matches_csv(items: list[CorpusEvaluationMatch], output_path: Path) -> None:
    """Пишет главную таблицу построчного сопоставления."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "проект",
                "project_id",
                "критерий",
                "критерий_код",
                "строка/диапазон",
                "строка_excel",
                "текст эталонной ошибки",
                "найденная ошибка",
                "найденная строка/диапазон",
                "finding_id",
                "checker",
                "тип совпадения",
                "засчитано",
                "score",
                "причина, почему засчитали",
            ],
        )
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "проект": item.project,
                    "project_id": item.project_id,
                    "критерий": item.label,
                    "критерий_код": item.criterion,
                    "строка/диапазон": item.gold_line_range,
                    "строка_excel": item.gold_row_number,
                    "текст эталонной ошибки": item.gold_text,
                    "найденная ошибка": item.found_text,
                    "найденная строка/диапазон": item.found_line_range,
                    "finding_id": item.found_finding_id or "",
                    "checker": item.found_checker or "",
                    "тип совпадения": item.match_type,
                    "засчитано": "да" if item.counted else "нет",
                    "score": item.match_score,
                    "причина, почему засчитали": item.reason,
                }
            )


def _write_false_negative_analysis_csv(items: list[FalseNegativeAnalysisItem], output_path: Path) -> None:
    """Пишет пропущенные ошибки с инженерной причиной пропуска."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "проект",
                "project_id",
                "критерий",
                "критерий_код",
                "строка/диапазон",
                "текст эталонной ошибки",
                "ближайшая находка",
                "nearest_checker",
                "nearest_match_type",
                "nearest_score",
                "reason_code",
                "причина пропуска",
                "следующий шаг",
            ],
        )
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "проект": item.project,
                    "project_id": item.project_id,
                    "критерий": item.label,
                    "критерий_код": item.criterion,
                    "строка/диапазон": item.gold_line_range,
                    "текст эталонной ошибки": item.gold_text,
                    "ближайшая находка": item.nearest_finding_id or "",
                    "nearest_checker": item.nearest_checker or "",
                    "nearest_match_type": item.nearest_match_type,
                    "nearest_score": item.nearest_score,
                    "reason_code": item.reason_code,
                    "причина пропуска": item.reason_label,
                    "следующий шаг": item.next_step,
                }
            )


def _write_detailed_false_positive_csv(items: list[PredictedCorpusItem], output_path: Path) -> None:
    """Пишет детальные ложные срабатывания, которые не сопоставились с эталоном."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "проект",
                "project_id",
                "критерий",
                "критерий_код",
                "найденная строка/диапазон",
                "найденная ошибка",
                "finding_id",
                "checker",
            ],
        )
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "проект": item.project,
                    "project_id": item.project_id,
                    "критерий": _criterion_label(item.criterion),
                    "критерий_код": item.criterion,
                    "найденная строка/диапазон": _format_prediction_range(item),
                    "найденная ошибка": item.found_text,
                    "finding_id": item.finding_id,
                    "checker": item.checker_name,
                }
            )


def _write_items_csv(items: list[CorpusEvaluationKey], output_path: Path) -> None:
    """Пишет ошибки сравнения: ложные пропуски или ложные срабатывания."""

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["project_id", "criterion", "label"])
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "project_id": item.project_id,
                    "criterion": item.criterion,
                    "label": _criterion_label(item.criterion),
                }
            )


def _normalize_project_name(value: str) -> str:
    """Нормализует имя проекта для сопоставления Excel и папок."""

    text = value.lower().replace("с", "c")
    text = re.sub(r"\.id_\d+.*$", "", text)
    text = re.sub(r"\(\d+\)", "", text)
    text = re.sub(r"[^a-zа-яё0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _project_tokens(value: str) -> list[str]:
    """Выделяет значимые токены имени проекта."""

    normalized = _normalize_project_name(value)
    return [token for token in normalized.split() if token not in {"master", "id"}]


def _criterion_label(criterion_value: str) -> str:
    """Возвращает русское название критерия."""

    try:
        return CRITERION_LABELS[Criterion(criterion_value)]
    except ValueError:
        return criterion_value


def _safe_ratio(numerator: int, denominator: int) -> float:
    """Делит без исключения на пустом наборе."""

    return round(numerator / denominator, 4) if denominator else 0.0


def _safe_money_ratio(cost: float, denominator: int) -> float | None:
    """Считает денежную метрику и явно возвращает None, если делить не на что."""

    return round(cost / denominator, 6) if denominator else None


def _f1(true_positive: int, false_positive: int, false_negative: int) -> float:
    """Считает F1 через precision и recall."""

    precision = _safe_ratio(true_positive, true_positive + false_positive)
    recall = _safe_ratio(true_positive, true_positive + false_negative)
    return round(2 * precision * recall / (precision + recall), 4) if precision + recall else 0.0


def _mean(values: list[float]) -> float:
    """Среднее значение для macro-метрик."""

    return round(sum(values) / len(values), 4) if values else 0.0
