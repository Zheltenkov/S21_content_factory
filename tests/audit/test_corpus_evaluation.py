from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook

from content_factory.audit.corpus_evaluation import (
    CorpusEvaluationMatch,
    _classify_false_negative,
    _split_gold_detail_cases,
    evaluate_corpus_report,
    write_corpus_evaluation,
)
from content_factory.audit.domain import (
    AuditReport,
    ContentUnit,
    Criterion,
    Evidence,
    Finding,
    ModelUsageSummary,
    RunSummary,
    Severity,
    TextLocation,
    Verdict,
)


def test_corpus_evaluation_matches_projects_and_computes_metrics(workspace_tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Проект", "SDP/ревью", "Проблема", "Детали", "Архив"])
    sheet.append(
        [
            "AP1-Go-T01",
            "",
            "неактуальная ссылка\nопечатки",
            "5 - Сломанная ссылка: https:/\\new.oprosso.net\n"
            "119 - Input operation без двоеточия\n"
            "122 - Input right operand без двоеточия",
            "",
        ]
    )
    sheet.append(["SQLB3", "", "несоответствие задания чек-листу", "В чеклисте ожидается другое условие.", ""])
    gold_path = workspace_tmp_path / "gold.xlsx"
    workbook.save(gold_path)

    report = AuditReport(
        summary=RunSummary(
            started_at=datetime.now(timezone.utc),
            input_path=str(workspace_tmp_path),
            findings_total=4,
            model_usage=ModelUsageSummary(calls_total=2, cache_hits=1, total_tokens=1200, cost_usd=0.24),
        ),
        units=[
            ContentUnit(
                unit_id="ap1_go_t01__abc",
                name="AP1_Go_T01.ID_1375359-master",
                root_path=workspace_tmp_path,
                relative_path="AP1_Go_T01.ID_1375359-master",
            ),
            ContentUnit(
                unit_id="sqlb3__abc",
                name="SQLB3_Retrieving_data.ID_574089-master (1)",
                root_path=workspace_tmp_path,
                relative_path="SQLB3_Retrieving_data.ID_574089-master (1)",
            ),
        ],
        entities=[],
        findings=[
            Finding(
                finding_id="f1",
                unit_id="ap1_go_t01__abc",
                branch=None,
                criterion=Criterion.LINKS,
                severity=Severity.MAJOR,
                verdict=Verdict.FAIL,
                confidence=0.9,
                quote="https:/\\new.oprosso.net",
                location=TextLocation(file_path="README.md", line_start=5, line_end=5),
                evidence=[Evidence(title="Ссылка", detail="Сломанная ссылка: https:/\\new.oprosso.net")],
                recommendation="Исправить ссылку.",
                checker_name="test",
            ),
            Finding(
                finding_id="f2",
                unit_id="ap1_go_t01__abc",
                branch=None,
                criterion=Criterion.READABILITY,
                severity=Severity.MINOR,
                verdict=Verdict.WARNING,
                confidence=0.9,
                quote="Input operation",
                location=TextLocation(file_path="README_RUS.md", line_start=119, line_end=119),
                evidence=[Evidence(title="Подпись", detail="Input operation без двоеточия")],
                recommendation="Добавить двоеточие.",
                checker_name="test",
            ),
            Finding(
                finding_id="f3",
                unit_id="ap1_go_t01__abc",
                branch=None,
                criterion=Criterion.READABILITY,
                severity=Severity.INFO,
                verdict=Verdict.UNKNOWN,
                confidence=0.4,
                quote="Лишняя редакторская находка",
                location=TextLocation(file_path="README.md", line_start=200, line_end=200),
                evidence=[Evidence(title="Редактура", detail="Этого нет в эталоне.")],
                recommendation="Проверить формулировку.",
                checker_name="test",
            ),
            Finding(
                finding_id="f4",
                unit_id="sqlb3__abc",
                branch=None,
                criterion=Criterion.CHECKLIST_ALIGNMENT,
                severity=Severity.MAJOR,
                verdict=Verdict.FAIL,
                confidence=0.9,
                quote="В чеклисте ожидается другое условие.",
                evidence=[Evidence(title="Чек-лист", detail="В чеклисте ожидается другое условие.")],
                recommendation="Синхронизировать чек-лист.",
                checker_name="test",
            ),
        ],
    )

    summary = evaluate_corpus_report(report, gold_path)

    assert summary.gold_total == 4
    assert summary.predicted_total == 3
    assert summary.true_positive == 3
    assert summary.false_positive == 0
    assert summary.false_negative == 1
    assert summary.precision == 1.0
    assert summary.recall == 0.75
    assert summary.f1_score == 0.8571
    assert summary.overview_precision == 1.0
    assert summary.overview_recall == 1.0
    assert summary.overview_f1_score == 1.0
    assert [item.match_type for item in summary.matches] == [
        "line_and_text",
        "line_and_text",
        "missed",
        "text_similarity",
    ]
    assert [item.counted for item in summary.matches] == [True, True, False, True]
    missed = next(item for item in summary.matches if "Input right operand" in item.gold_text)
    assert missed.found_text == ""
    assert "не найдено подходящей ошибки" in missed.reason
    assert summary.detailed_false_positive_items == []
    assert summary.actionable_metrics is not None
    assert summary.actionable_metrics.predicted_total == 3
    assert summary.actionable_metrics.true_positive == 3
    assert summary.actionable_metrics.precision == 1.0
    assert summary.cost_quality is not None
    assert summary.cost_quality.cost_per_gold_true_positive == 0.08
    assert summary.cost_quality.cost_per_prediction == 0.06
    assert summary.checker_metrics[0].slice_name == "test"
    assert summary.checker_group_metrics[0].slice_name == "other"
    assert summary.false_negative_reason_counts == {"deterministic_possible": 1}
    assert summary.false_negative_analysis[0].reason_code == "deterministic_possible"

    output_dir = workspace_tmp_path / "evaluation"
    write_corpus_evaluation(summary, output_dir)
    main_csv = (output_dir / "corpus_evaluation_main.csv").read_text(encoding="utf-8-sig")
    assert "проект,project_id,критерий" in main_csv
    assert "текст эталонной ошибки" in main_csv
    assert "найденная ошибка" in main_csv
    assert "тип совпадения" in main_csv
    assert "причина, почему засчитали" in main_csv
    assert (output_dir / "corpus_evaluation_overview_by_criterion.csv").exists()
    assert (output_dir / "corpus_evaluation_by_checker.csv").exists()
    assert (output_dir / "corpus_evaluation_by_checker_group.csv").exists()
    assert (output_dir / "corpus_false_negative_analysis.csv").exists()


def test_corpus_evaluation_splits_only_real_gold_defects() -> None:
    cases = _split_gold_detail_cases(
        "Следов выполнения команды whoami в дампе нет.\n"
        "\n"
        "Предложение по решению:\n"
        "Убрать из чек-листа указание на команду whoami\n"
        "Аргументы и исследования на тему:\n"
        "Скрины в треде"
    )

    assert len(cases) == 1
    assert cases[0]["text"] == "Следов выполнения команды whoami в дампе нет."
    assert cases[0]["line_start"] is None


def test_corpus_evaluation_does_not_treat_review_list_number_as_source_line() -> None:
    cases = _split_gold_detail_cases(
        "1. Google C++ style guide не является стандартным для C.\n"
        "2. Лучше начинать изучение языка с C99.\n"
        "101–110 - Все 10 пунктов пронумерованы как 1)"
    )

    assert [case["text"] for case in cases] == [
        "Google C++ style guide не является стандартным для C.",
        "Лучше начинать изучение языка с C99.",
        "101–110 - Все 10 пунктов пронумерованы как 1)",
    ]
    assert cases[0]["line_start"] is None
    assert cases[1]["line_start"] is None
    assert cases[2]["line_start"] == 101
    assert cases[2]["line_end"] == 110


def test_false_negative_classifier_does_not_treat_weak_criterion_match_as_near_miss() -> None:
    match = CorpusEvaluationMatch(
        project="AP1_Go_T01",
        project_id="ap1",
        criterion="readability",
        label="Грамотность и читаемость текста",
        gold_row_number=2,
        gold_line_range="119",
        gold_text="119 - Input operation без двоеточия",
        found_finding_id="fnd_other",
        found_checker="tech_freshness_checker",
        found_line_range="README.md:50",
        found_text="Нерелевантная находка того же критерия",
        match_type="criterion_only",
        match_score=0.2,
        counted=False,
        reason="Совпал только критерий.",
    )

    reason_code, _label, _next_step = _classify_false_negative(match)

    assert reason_code == "deterministic_possible"


def test_corpus_evaluation_matches_missing_artifact_command_output(workspace_tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Проект", "SDP/ревью", "Проблема", "Детали", "Архив"])
    sheet.append(
        [
            "CbS4",
            "",
            "ошибка в чек-листе",
            "Следов выполнения команды whoami через Reverse Shell в прилагаемом файле дампа (pcapng) нет.",
            "",
        ]
    )
    gold_path = workspace_tmp_path / "gold.xlsx"
    workbook.save(gold_path)

    report = AuditReport(
        summary=RunSummary(started_at=datetime.now(timezone.utc), input_path=str(workspace_tmp_path)),
        units=[
            ContentUnit(
                unit_id="cbs4__abc",
                name="CbS4_Networking_basics_Part_4.ID_1521091-master",
                root_path=workspace_tmp_path,
                relative_path="CbS4_Networking_basics_Part_4.ID_1521091-master",
            )
        ],
        entities=[],
        findings=[
            Finding(
                finding_id="f1",
                unit_id="cbs4__abc",
                branch=None,
                criterion=Criterion.CHECKLIST_ALIGNMENT,
                severity=Severity.MAJOR,
                verdict=Verdict.WARNING,
                confidence=0.88,
                quote="вывод whoami не найден в src/task_2/some_troubled_traffic.pcapng",
                location=TextLocation(file_path="check-list.yml"),
                evidence=[Evidence(title="Артефакт", detail="Ожидаемый вывод команды в pcapng не найден.")],
                recommendation="Проверить артефакт или чек-лист.",
                checker_name="checklist_checker",
            )
        ],
    )

    summary = evaluate_corpus_report(report, gold_path)

    assert summary.true_positive == 1
    assert summary.matches[0].match_type == "artifact_missing_signal"


def test_corpus_evaluation_ignores_diagnostic_predictions_in_strict_metric(workspace_tmp_path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Проект", "SDP/ревью", "Проблема", "Детали", "Архив"])
    sheet.append(["Unit", "", "неактуальная ссылка", "5 - Сломанная ссылка.", ""])
    sheet.append(["Unit", "", "ошибка в задании", "10 - Отсутствует конкретный датасет.", ""])
    gold_path = workspace_tmp_path / "gold.xlsx"
    workbook.save(gold_path)

    unit = ContentUnit(
        unit_id="unit__abc",
        name="Unit.ID_1-master",
        root_path=workspace_tmp_path,
        relative_path="Unit.ID_1-master",
    )
    report = AuditReport(
        summary=RunSummary(started_at=datetime.now(timezone.utc), input_path=str(workspace_tmp_path)),
        units=[unit],
        entities=[],
        findings=[
            Finding(
                finding_id="link-unknown",
                unit_id=unit.unit_id,
                branch=None,
                criterion=Criterion.ACTUALITY,
                severity=Severity.INFO,
                verdict=Verdict.UNKNOWN,
                confidence=0.5,
                quote="http://example.test",
                location=TextLocation(file_path="README.md", line_start=5, line_end=5),
                evidence=[Evidence(title="Сеть отключена", detail="Ссылка не проверялась.")],
                recommendation="Запустить проверку с сетью.",
                checker_name="link_checker",
            ),
            Finding(
                finding_id="generic-resource",
                unit_id=unit.unit_id,
                branch=None,
                criterion=Criterion.CORRECTNESS,
                severity=Severity.MAJOR,
                verdict=Verdict.WARNING,
                confidence=0.78,
                quote="dataset",
                location=TextLocation(file_path="README.md", line_start=10, line_end=10),
                evidence=[Evidence(title="Локальный ресурс", detail="В тексте нужен локальный ресурс.")],
                recommendation="Указать конкретный файл/ссылку на ресурс.",
                checker_name="resource_availability_checker",
                extra={"issue_type": "resource_without_artifact"},
            ),
        ],
    )

    summary = evaluate_corpus_report(report, gold_path)

    assert summary.predicted_total == 0
    assert summary.true_positive == 0
    assert summary.false_negative == 2
